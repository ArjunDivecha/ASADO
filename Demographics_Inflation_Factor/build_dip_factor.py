"""
build_dip_factor.py
===================
Reproducible pipeline: builds the Demographic Inflation Pressure (DIP) factor
for the ASADO 34-country universe from UN WPP 2024 cohort data, using the
coefficients from Juselius & Takats (2018), BIS WP 722.

Run from the Demographics_Inflation_Factor folder:
    python build_dip_factor.py

Outputs (created in current directory):
    beta_coeffs.csv             17 cohort coefficients beta_1k
    dip_panel_long.csv          long format: ISO3, Year, DIP_abs, DIP_rel, chg_5y, chg_10y
    dip_rel_wide.csv            wide: years x countries, cross-country relative DIP
    dip_abs_wide.csv            wide: years x countries, absolute DIP
    snapshot_2025_2050.csv      forward snapshot at 2025/27/30/35/40/45/50
    ASADO_DIP_factor.xlsx       multi-sheet workbook with conditional formatting
    us_validation.png           US DIP_abs vs paper Graph 4
    crosscountry_ranking.png    bar charts: 2025/2035/2050 cross-country ranking
    selected_countries.png      time series for USA, JPN, DEU, GBR, KOR, CHN, IND, BRA, MEX

Dependencies: pandas, numpy, matplotlib, openpyxl, requests
"""
import os
import gzip
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ---------------------------------------------------------------------------
# 1. CONFIGURATION
# ---------------------------------------------------------------------------

WPP_URL = ("https://population.un.org/wpp/assets/Excel%20Files/"
           "1_Indicator%20%28Standard%29/CSV_FILES/"
           "WPP2024_PopulationByAge5GroupSex_Medium.csv.gz")
WPP_LOCAL_GZ = "wpp2024_age5.csv.gz"
WPP_LOCAL_CSV = "wpp2024_age5.csv"

# ASADO 34-country universe
ASADO_22 = ["AUT","AUS","BEL","CAN","CHE","DEU","DNK","ESP","FIN","FRA",
            "GBR","GRC","IRL","ITA","JPN","KOR","NLD","NOR","NZL","PRT","SWE","USA"]
ASADO_EM = ["CHN","IND","BRA","MEX","RUS","ZAF","IDN","THA","MYS","POL","TUR","CHL"]
ISO_LIST = ASADO_22 + ASADO_EM

# Juselius & Takats (2018) Table 1 Model 2 polynomial coefficients (with time FE).
# Displayed values in the paper are scaled by 10^(p-1); we recover the raw gamma_p.
GAMMA = {
    1: 0.74  / 1,        #  0.7400
    2: -1.83 / 10,       # -0.1830
    3: 1.68  / 100,      #  0.0168
    4: -0.50 / 1000,     # -0.0005
}
K_COHORTS = 17  # 0-4, 5-9, ..., 75-79, 80+

# ---------------------------------------------------------------------------
# 2. RECOVER 17 COHORT COEFFICIENTS beta_1k FROM POLYNOMIAL
# ---------------------------------------------------------------------------

def recover_beta_1k():
    """beta_1k = gamma_0 + sum_p gamma_p * k^p, with sum_k beta_1k = 0 (Eq B5)."""
    gamma_0 = -sum(GAMMA[p] * sum(k**p for k in range(1, K_COHORTS + 1)) / K_COHORTS
                   for p in range(1, 5))
    beta = np.array([
        gamma_0 + sum(GAMMA[p] * k**p for p in range(1, 5))
        for k in range(1, K_COHORTS + 1)
    ])
    labels = [f"{(k-1)*5}-{(k-1)*5+4}" if k < K_COHORTS else "80+"
              for k in range(1, K_COHORTS + 1)]
    return pd.DataFrame({"k": range(1, K_COHORTS + 1),
                         "cohort": labels,
                         "beta_1k": beta})

# ---------------------------------------------------------------------------
# 3. DOWNLOAD AND PARSE WPP 2024
# ---------------------------------------------------------------------------

def ensure_wpp():
    """Download WPP 2024 5-year age group CSV if not present locally."""
    if Path(WPP_LOCAL_CSV).exists():
        return
    if not Path(WPP_LOCAL_GZ).exists():
        print(f"Downloading {WPP_URL} (~30 MB) ...")
        import requests
        r = requests.get(WPP_URL, stream=True)
        r.raise_for_status()
        with open(WPP_LOCAL_GZ, "wb") as f:
            shutil.copyfileobj(r.raw, f)
    print(f"Decompressing {WPP_LOCAL_GZ} ...")
    with gzip.open(WPP_LOCAL_GZ, "rb") as src, open(WPP_LOCAL_CSV, "wb") as dst:
        shutil.copyfileobj(src, dst)

def load_cohort_shares():
    """Read WPP CSV, filter to ASADO universe, aggregate to 17 cohorts, compute shares."""
    chunks = []
    cols = ["ISO3_code","Location","Variant","Time","AgeGrp","AgeGrpStart","PopTotal"]
    for chunk in pd.read_csv(WPP_LOCAL_CSV, chunksize=200_000,
                             low_memory=False, usecols=cols):
        chunk = chunk[chunk["ISO3_code"].isin(ISO_LIST)
                      & (chunk["Variant"] == "Medium")]
        if len(chunk):
            chunks.append(chunk)
    df = pd.concat(chunks, ignore_index=True)
    df["Time"] = pd.to_numeric(df["Time"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["Time"])

    # Map age-start to cohort k=1..17 (80+ is open-ended)
    df["k"] = df["AgeGrpStart"].apply(lambda a: 17 if a >= 80 else int(a // 5) + 1)
    panel = df.groupby(["ISO3_code","Location","Time","k"], as_index=False)["PopTotal"].sum()

    wide = panel.pivot_table(index=["ISO3_code","Location","Time"],
                             columns="k", values="PopTotal").reset_index()
    wide.columns = ["ISO3","Location","Year"] + [f"N{k}" for k in range(1, 18)]
    N_cols = [f"N{k}" for k in range(1, 18)]
    wide["N_total"] = wide[N_cols].sum(axis=1)
    for k in range(1, 18):
        wide[f"n{k}"] = wide[f"N{k}"] / wide["N_total"]
    return wide

# ---------------------------------------------------------------------------
# 4. COMPUTE DIP (absolute and cross-country relative)
# ---------------------------------------------------------------------------

def compute_dip(panel, beta):
    """Add DIP_abs (absolute) and DIP_rel (deviation from cross-country mean) columns."""
    share_cols = [f"n{k}" for k in range(1, 18)]
    yearly_mean = panel.groupby("Year")[share_cols].mean().reset_index()
    yearly_mean.columns = ["Year"] + [f"{c}_mean" for c in share_cols]
    panel = panel.merge(yearly_mean, on="Year")
    dev_cols = []
    for k in range(1, 18):
        dev_col = f"dev{k}"
        panel[dev_col] = panel[f"n{k}"] - panel[f"n{k}_mean"]
        dev_cols.append(dev_col)
    panel["DIP_abs"] = (panel[share_cols].values * beta).sum(axis=1) * 100
    panel["DIP_rel"] = (panel[dev_cols].values * beta).sum(axis=1) * 100
    panel = panel.sort_values(["ISO3","Year"]).reset_index(drop=True)
    panel["DIP_rel_chg_5y"]  = panel.groupby("ISO3")["DIP_rel"].diff(5)
    panel["DIP_rel_chg_10y"] = panel.groupby("ISO3")["DIP_rel"].diff(10)
    return panel

# ---------------------------------------------------------------------------
# 5. CHARTS
# ---------------------------------------------------------------------------

def make_charts(panel):
    # US validation
    us = panel[panel["ISO3"] == "USA"].sort_values("Year").copy()
    hist = us[(us["Year"] >= 1950) & (us["Year"] <= 2020)]
    us["DIP_abs_norm"] = us["DIP_abs"] - hist["DIP_abs"].mean()
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(us["Year"], us["DIP_abs_norm"], color="#d62728", lw=2,
            label="Demographic inflation pressure (centered)")
    ax.axhline(0, color="gray", lw=0.5)
    ax.axvspan(2024, 2050, alpha=0.08, color="blue", label="WPP 2024 projection")
    ax.set_title("US Demographic Inflation Pressure — Juselius & Takats (2018) replication\n"
                 "(Table 1 Model 2 coefficients applied to WPP 2024)", fontsize=11)
    ax.set_xlabel("Year"); ax.set_ylabel("Pressure (pp inflation)")
    ax.legend(loc="upper right"); ax.grid(alpha=0.3)
    plt.tight_layout(); plt.savefig("us_validation.png", dpi=140, bbox_inches="tight"); plt.close()

    # Cross-country ranking 2025/2035/2050
    fig, axes = plt.subplots(1, 3, figsize=(15, 6), sharey=True)
    for ax, yr in zip(axes, [2025, 2035, 2050]):
        snap = panel[panel["Year"] == yr][["ISO3","DIP_rel"]].sort_values("DIP_rel")
        colors = ["#1f77b4" if iso in ASADO_22 else "#ff7f0e" for iso in snap["ISO3"]]
        ax.barh(snap["ISO3"], snap["DIP_rel"], color=colors, edgecolor="black", lw=0.3)
        ax.axvline(0, color="black", lw=0.5)
        ax.set_title(str(yr)); ax.set_xlabel("Cross-country DIP (pp)")
        ax.grid(alpha=0.3, axis="x")
    axes[0].set_ylabel("Country")
    fig.suptitle("Demographic Inflation Pressure — cross-country relative score\n"
                 "(blue = paper's advanced economies, orange = EM addition)", fontsize=11)
    plt.tight_layout(); plt.savefig("crosscountry_ranking.png", dpi=140, bbox_inches="tight"); plt.close()

    # Selected countries time series
    fig, axes = plt.subplots(3, 3, figsize=(15, 11), sharex=True, sharey=True)
    for ax, iso in zip(axes.flat, ["USA","JPN","DEU","GBR","KOR","CHN","IND","BRA","MEX"]):
        c = panel[panel["ISO3"] == iso].sort_values("Year")
        ax.plot(c["Year"], c["DIP_rel"], color="#1f77b4", lw=1.5)
        ax.axhline(0, color="gray", lw=0.5)
        ax.axvspan(2024, 2100, alpha=0.08, color="blue")
        ax.set_title(iso, fontsize=11); ax.grid(alpha=0.3); ax.set_xlim(1950, 2060)
    for ax in axes[-1,:]: ax.set_xlabel("Year")
    for ax in axes[:,0]:  ax.set_ylabel("DIP_rel (pp)")
    fig.suptitle("Cross-country Demographic Inflation Pressure — selected ASADO countries\n"
                 "(positive = relatively more inflationary demographics than panel average)",
                 fontsize=12)
    plt.tight_layout(); plt.savefig("selected_countries.png", dpi=140, bbox_inches="tight"); plt.close()

# ---------------------------------------------------------------------------
# 6. EXCEL WORKBOOK
# ---------------------------------------------------------------------------

def build_workbook(panel, beta_df):
    rel_wide = panel.pivot(index="Year", columns="ISO3", values="DIP_rel")[ISO_LIST].reset_index()
    abs_wide = panel.pivot(index="Year", columns="ISO3", values="DIP_abs")[ISO_LIST].reset_index()
    snap_years = [2025, 2027, 2030, 2035, 2040, 2045, 2050]
    snap = panel[panel["Year"].isin(snap_years)].pivot(
        index="ISO3", columns="Year", values="DIP_rel").loc[ISO_LIST].reset_index()

    wb = Workbook()
    bold = Font(bold=True)
    color_rule = ColorScaleRule(start_type="min", start_color="5B9BD5",
                                mid_type="num", mid_value=0, mid_color="FFFFFF",
                                end_type="max", end_color="E63946")

    # Coefficients
    ws = wb.active; ws.title = "Coefficients"
    ws.append(["k","Cohort","beta_1k"])
    for r in beta_df.itertuples(index=False):
        ws.append([r.k, r.cohort, round(r.beta_1k, 6)])
    for col in "ABC":
        ws.column_dimensions[col].width = 18
    for c in ws[1]: c.font = bold

    # DIP_rel_wide
    ws = wb.create_sheet("DIP_rel_wide")
    ws.append(list(rel_wide.columns))
    for r in rel_wide.round(4).itertuples(index=False): ws.append(list(r))
    for c in ws[1]: c.font = bold
    rng = f"B2:{get_column_letter(len(rel_wide.columns))}{len(rel_wide)+1}"
    ws.conditional_formatting.add(rng, color_rule); ws.freeze_panes = "B2"

    # DIP_abs_wide
    ws = wb.create_sheet("DIP_abs_wide")
    ws.append(list(abs_wide.columns))
    for r in abs_wide.round(4).itertuples(index=False): ws.append(list(r))
    for c in ws[1]: c.font = bold; ws.freeze_panes = "B2"

    # Snapshot
    ws = wb.create_sheet("Snapshot_2025_2050")
    ws.append(["ISO3"] + [str(y) for y in snap_years] + ["Delta 2025->2050"])
    for r in snap.itertuples(index=False):
        vals = list(r); delta = vals[-1] - vals[1]
        ws.append([vals[0]] + [round(v,3) for v in vals[1:]] + [round(delta,3)])
    for c in ws[1]: c.font = bold
    rng = f"B2:{get_column_letter(len(snap_years)+2)}{len(snap)+1}"
    ws.conditional_formatting.add(rng, color_rule)

    # Panel long
    ws = wb.create_sheet("Panel_long")
    out_cols = ["ISO3","Location","Year","DIP_abs","DIP_rel","DIP_rel_chg_5y","DIP_rel_chg_10y"]
    ws.append(out_cols)
    for r in panel[out_cols].round(4).itertuples(index=False): ws.append(list(r))
    for c in ws[1]: c.font = bold; ws.freeze_panes = "A2"

    wb.save("ASADO_DIP_factor.xlsx")

# ---------------------------------------------------------------------------
# 7. MAIN
# ---------------------------------------------------------------------------

def main():
    print("Step 1: cohort coefficients")
    beta_df = recover_beta_1k()
    beta_df.to_csv("beta_coeffs.csv", index=False)
    beta = beta_df["beta_1k"].values
    print(f"  17 cohort betas, sum = {beta.sum():.6f} (should be ~0)")

    print("Step 2: WPP 2024 data")
    ensure_wpp()
    panel_shares = load_cohort_shares()
    print(f"  {panel_shares['ISO3'].nunique()} countries x {panel_shares['Year'].nunique()} years")

    print("Step 3: compute DIP")
    panel = compute_dip(panel_shares, beta)

    print("Step 4: write CSV outputs")
    panel[["ISO3","Location","Year","DIP_abs","DIP_rel","DIP_rel_chg_5y","DIP_rel_chg_10y"]] \
        .round(4).to_csv("dip_panel_long.csv", index=False)
    panel.pivot(index="Year", columns="ISO3", values="DIP_rel")[ISO_LIST] \
        .round(4).to_csv("dip_rel_wide.csv")
    panel.pivot(index="Year", columns="ISO3", values="DIP_abs")[ISO_LIST] \
        .round(4).to_csv("dip_abs_wide.csv")

    snap_years = [2025, 2027, 2030, 2035, 2040, 2045, 2050]
    snap = panel[panel["Year"].isin(snap_years)].pivot(
        index="ISO3", columns="Year", values="DIP_rel")[snap_years].loc[ISO_LIST]
    snap["Delta_2025_2050"] = snap[2050] - snap[2025]
    snap.round(4).to_csv("snapshot_2025_2050.csv")

    print("Step 5: charts")
    make_charts(panel)

    print("Step 6: Excel workbook")
    build_workbook(panel, beta_df)

    print("\nDone. Outputs in current directory.")

if __name__ == "__main__":
    main()
