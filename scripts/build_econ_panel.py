#!/usr/bin/env python3
"""
=============================================================================
SCRIPT NAME: scripts/build_econ_panel.py
=============================================================================

INPUT FILES:
- Data/asado.duckdb :: unified_panel — all 146 Econ variables, tidy long format
  (date, country, value, variable, source)

OUTPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/Data/work/econ/Econ.xlsx
    Wide monthly workbook consumed by Step Two Econ Create Tidy.py.
    Sheet 1: INDEX — sheet_id, sheet_name, variable, source, source_table,
                     first_date, last_date, n_dates, n_countries_with_data
    Sheets 2..147: one per variable, rows = first-of-month dates,
                   column A = "date", columns B..AH = 34 T2 country names
                   in canonical alphabetical order.
- Data/processed/econ_workbook_panel.parquet — local snapshot (wide form,
    stacked, backed up before overwrite).

VERSION: 1.0
LAST UPDATED: 2026-04-29
AUTHOR: Arjun Divecha (with Claude)

DESCRIPTION:
Reads all 146 Econ factor variables from ASADO's unified_panel DuckDB table,
pivots each to wide format (dates × countries), and writes Econ.xlsx directly
to the T2 Econ directory — the final destination consumed by Step Two Econ.

Replaces the previous manual process of exporting Econ.xlsx from Bloomberg /
external data sources. All Econ data is now produced within ASADO.

DEPENDENCIES:
- duckdb, pandas, openpyxl, pyarrow

USAGE:
  python scripts/build_econ_panel.py              # normal run
  python scripts/build_econ_panel.py --dry-run    # preview, no writes
  python scripts/build_econ_panel.py --check      # report on existing output

NOTES:
- Sheet names are truncated to 31 chars (Excel limit). The INDEX sheet
  contains the full variable name for all 14 truncated cases.
- Variables missing from unified_panel are written as empty sheets with a
  warning — does NOT abort the run.
- Backup: Data/backups/{timestamp}/econ_workbook_panel.parquet before overwrite.
=============================================================================
"""

from __future__ import annotations

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import duckdb
import pandas as pd

# Econ redundancy drop-list (single source of truth for the prune rule).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from econ_drop_list import should_drop  # noqa: E402

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"
PROCESSED_DIR = DATA_DIR / "processed"
BACKUPS_DIR = DATA_DIR / "backups"
DUCKDB_PATH = DATA_DIR / "asado.duckdb"

OUTPUT_XLSX = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/Data/work/econ/Econ.xlsx"
)
LOCAL_PARQUET = PROCESSED_DIR / "econ_workbook_panel.parquet"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ── 34 T2 countries in canonical column order (matches existing Econ.xlsx) ──
T2_COUNTRIES = [
    "Australia", "Brazil", "Canada", "Chile", "ChinaA", "ChinaH", "Denmark",
    "France", "Germany", "Hong Kong", "India", "Indonesia", "Italy", "Japan",
    "Korea", "Malaysia", "Mexico", "NASDAQ", "Netherlands", "Philippines",
    "Poland", "Saudi Arabia", "Singapore", "South Africa", "Spain", "Sweden",
    "Switzerland", "Taiwan", "Thailand", "Turkey", "U.K.", "U.S.",
    "US SmallCap", "Vietnam",
]

# ── Sheet name → full variable name (146 entries, order = INDEX order) ──────
# Sheet names are Excel-safe (≤31 chars). Where the variable name is longer,
# the INDEX-derived short name is kept for downstream compatibility.
SHEETS: list[tuple[str, str]] = [
    ("MS_Country_ETF_AUM_USD",          "MS_Country_ETF_AUM_USD"),
    ("MS_Country_ETF_NetFlow_USD",       "MS_Country_ETF_NetFlow_USD"),
    ("MS_ETF_Creation_Fee_USD",          "MS_ETF_Creation_Fee_USD"),
    ("MS_ETF_Creation_Unit_Size__d549",  "MS_ETF_Creation_Unit_Size_Shares"),
    ("MS_ETF_NetCreation_Shares",        "MS_ETF_NetCreation_Shares"),
    ("MS_ETF_NetFlow_to_MarketCap",      "MS_ETF_NetFlow_to_MarketCap"),
    ("MS_ETF_Redemption_Fee_USD",        "MS_ETF_Redemption_Fee_USD"),
    ("MS_Passive_AUM_to_MarketCap",      "MS_Passive_AUM_to_MarketCap"),
    ("MS_Passive_Flow_Distortion",       "MS_Passive_Flow_Distortion"),
    ("BIS_DSR_Private",                  "BIS_DSR_Private"),
    ("BIS_Policy_Rate",                  "BIS_Policy_Rate"),
    ("ECB_FX_AUD_EUR",                   "ECB_FX_AUD_EUR"),
    ("ECB_FX_BRL_EUR",                   "ECB_FX_BRL_EUR"),
    ("ECB_FX_CAD_EUR",                   "ECB_FX_CAD_EUR"),
    ("ECB_FX_CHF_EUR",                   "ECB_FX_CHF_EUR"),
    ("ECB_FX_DKK_EUR",                   "ECB_FX_DKK_EUR"),
    ("ECB_FX_EUR_EUR",                   "ECB_FX_EUR_EUR"),
    ("ECB_FX_GBP_EUR",                   "ECB_FX_GBP_EUR"),
    ("ECB_FX_HKD_EUR",                   "ECB_FX_HKD_EUR"),
    ("ECB_FX_IDR_EUR",                   "ECB_FX_IDR_EUR"),
    ("ECB_FX_INR_EUR",                   "ECB_FX_INR_EUR"),
    ("ECB_FX_JPY_EUR",                   "ECB_FX_JPY_EUR"),
    ("ECB_FX_KRW_EUR",                   "ECB_FX_KRW_EUR"),
    ("ECB_FX_MXN_EUR",                   "ECB_FX_MXN_EUR"),
    ("ECB_FX_MYR_EUR",                   "ECB_FX_MYR_EUR"),
    ("ECB_FX_PHP_EUR",                   "ECB_FX_PHP_EUR"),
    ("ECB_FX_PLN_EUR",                   "ECB_FX_PLN_EUR"),
    ("ECB_FX_SEK_EUR",                   "ECB_FX_SEK_EUR"),
    ("ECB_FX_SGD_EUR",                   "ECB_FX_SGD_EUR"),
    ("ECB_FX_THB_EUR",                   "ECB_FX_THB_EUR"),
    ("ECB_FX_TRY_EUR",                   "ECB_FX_TRY_EUR"),
    ("ECB_FX_TWD_EUR",                   "ECB_FX_TWD_EUR"),
    ("ECB_FX_USD_EUR",                   "ECB_FX_USD_EUR"),
    ("ECB_FX_ZAR_EUR",                   "ECB_FX_ZAR_EUR"),
    ("EIA_Petroleum_Consumption_TBPD",   "EIA_Petroleum_Consumption_TBPD"),
    ("FAO_AgExport_GDP_Share",           "FAO_AgExport_GDP_Share"),
    ("FAO_Import_Dependency",            "FAO_Import_Dependency"),
    ("FAO_Self_Sufficiency",             "FAO_Self_Sufficiency"),
    ("FAO_Terms_of_Trade",               "FAO_Terms_of_Trade"),
    ("FAO_Trade_Openness",               "FAO_Trade_Openness"),
    ("FRED_HY_OAS",                      "FRED_HY_OAS"),
    ("FRED_USD_Broad_Index",             "FRED_USD_Broad_Index"),
    ("FRED_UST_10Y",                     "FRED_UST_10Y"),
    ("FRED_UST_2Y",                      "FRED_UST_2Y"),
    ("FRED_VIX",                         "FRED_VIX"),
    ("FRED_Yield_Curve_10Y2Y",           "FRED_Yield_Curve_10Y2Y"),
    ("ILO_LFP_Rate",                     "ILO_LFP_Rate"),
    ("ILO_Unemployment_Rate",            "ILO_Unemployment_Rate"),
    ("NDGAIN_Readiness",                 "NDGAIN_Readiness"),
    ("NDGAIN_Score",                     "NDGAIN_Score"),
    ("NDGAIN_Vulnerability",             "NDGAIN_Vulnerability"),
    ("OECD_BCI",                         "OECD_BCI"),
    ("OECD_CCI",                         "OECD_CCI"),
    ("OFAC_Sanctioned",                  "OFAC_Sanctioned"),
    ("OFAC_Sanctions_Count",             "OFAC_Sanctions_Count"),
    ("UNDP_GDI",                         "UNDP_GDI"),
    ("UNDP_GII",                         "UNDP_GII"),
    ("UNDP_HDI",                         "UNDP_HDI"),
    ("UNDP_IHDI",                        "UNDP_IHDI"),
    ("BIS_Credit_GDP_Gap",               "BIS_Credit_GDP_Gap"),
    ("BIS_Property_Price",               "BIS_Property_Price"),
    ("BIS_REER",                         "BIS_REER"),
    ("EPU",                              "EPU"),
    ("GPR",                              "GPR"),
    ("Global_GPR",                       "Global_GPR"),
    ("Global_GPR_Act",                   "Global_GPR_Act"),
    ("Global_GPR_Threat",                "Global_GPR_Threat"),
    ("OECD_CLI",                         "OECD_CLI"),
    ("WB_CO2_Per_Capita",                "WB_CO2_Per_Capita"),
    ("WB_Control_Corruption",            "WB_Control_Corruption"),
    ("WB_Current_Account_GDP",           "WB_Current_Account_GDP"),
    ("WB_Domestic_Credit_GDP",           "WB_Domestic_Credit_GDP"),
    ("WB_External_Debt_GNI",             "WB_External_Debt_GNI"),
    ("WB_FDI_Inflows_GDP",               "WB_FDI_Inflows_GDP"),
    ("WB_FX_Reserves",                   "WB_FX_Reserves"),
    ("WB_Female_LFP",                    "WB_Female_LFP"),
    ("WB_Female_Labor_Share",            "WB_Female_Labor_Share"),
    ("WB_GDP_Growth_Real",               "WB_GDP_Growth_Real"),
    ("WB_Govt_Debt_GDP",                 "WB_Govt_Debt_GDP"),
    ("WB_Govt_Effectiveness",            "WB_Govt_Effectiveness"),
    ("WB_Import_Cover_Months",           "WB_Import_Cover_Months"),
    ("WB_Inflation_CPI",                 "WB_Inflation_CPI"),
    ("WB_Labor_Force",                   "WB_Labor_Force"),
    ("WB_Market_Cap_GDP",                "WB_Market_Cap_GDP"),
    ("WB_OldAge_Dependency",             "WB_OldAge_Dependency"),
    ("WB_Political_Stability",           "WB_Political_Stability"),
    ("WB_Population",                    "WB_Population"),
    ("WB_Population_Growth",             "WB_Population_Growth"),
    ("WB_Regulatory_Quality",            "WB_Regulatory_Quality"),
    ("WB_Renewable_Energy_Share",        "WB_Renewable_Energy_Share"),
    ("WB_Rule_of_Law",                   "WB_Rule_of_Law"),
    ("WB_Trade_Openness",                "WB_Trade_Openness"),
    ("WB_Unemployment",                  "WB_Unemployment"),
    ("WB_Voice_Accountability",          "WB_Voice_Accountability"),
    ("IMF_BOP_Current_Account",          "IMF_BOP_Current_Account"),
    ("IMF_BOP_Direct_Investment_Net",    "IMF_BOP_Direct_Investment_Net"),
    ("IMF_BOP_Financial_Account_Bal",    "IMF_BOP_Financial_Account_Bal"),
    ("IMF_BOP_Portfolio_Investme_10da",  "IMF_BOP_Portfolio_Investment_Net"),
    ("IMF_CPI_Index",                    "IMF_CPI_Index"),
    ("IMF_CPI_Inflation_YoY",            "IMF_CPI_Inflation_YoY"),
    ("IMF_XRate_LCU_per_USD",            "IMF_XRate_LCU_per_USD"),
    ("IMF_Export_Price_Index",           "IMF_Export_Price_Index"),
    ("IMF_Exports_USD",                  "IMF_Exports_USD"),
    ("IMF_Exports_YoY",                  "IMF_Exports_YoY"),
    ("IMF_Import_Price_Index",           "IMF_Import_Price_Index"),
    ("IMF_Imports_USD",                  "IMF_Imports_USD"),
    ("IMF_Imports_YoY",                  "IMF_Imports_YoY"),
    ("IMF_Trade_Balance_USD",            "IMF_Trade_Balance_USD"),
    ("IMF_Trade_Openness_USD",           "IMF_Trade_Openness_USD"),
    ("IMF_Employment_Index",             "IMF_Employment_Index"),
    ("IMF_Discount_Rate",                "IMF_Discount_Rate"),
    ("IMF_Govt_Bond_Yield",              "IMF_Govt_Bond_Yield"),
    ("IMF_Money_Market_Rate",            "IMF_Money_Market_Rate"),
    ("IMF_TBill_Rate",                   "IMF_TBill_Rate"),
    ("IMF_WEO_CA_GDP",                   "IMF_WEO_CA_GDP"),
    ("IMF_WEO_Debt_GDP",                 "IMF_WEO_Debt_GDP"),
    ("IMF_WEO_GDP_Growth",               "IMF_WEO_GDP_Growth"),
    ("IMF_WEO_Inflation",                "IMF_WEO_Inflation"),
    ("IMF_WEO_Population",               "IMF_WEO_Population"),
    ("IMF_WEO_Unemployment",             "IMF_WEO_Unemployment"),
    ("MS_Bank_Capital_Adequacy",         "MS_Bank_Capital_Adequacy"),
    ("MS_Bank_Liquidity_Coverage_9b13",  "MS_Bank_Liquidity_Coverage_Ratio"),
    ("MS_Bank_Liquidity_Ratio",          "MS_Bank_Liquidity_Ratio"),
    ("MS_Bank_Net_Stable_Funding_bd7e",  "MS_Bank_Net_Stable_Funding_Ratio"),
    ("MS_NPL_Net_Provisions_to_C_3fee",  "MS_NPL_Net_Provisions_to_Capital_Pct"),
    ("MS_NPL_Ratio",                     "MS_NPL_Ratio"),
    ("MS_CentralBank_BalanceSheet_GDP",  "MS_CentralBank_BalanceSheet_GDP"),
    ("MS_CentralBank_Claims_on_G_0cda",  "MS_CentralBank_Claims_on_Government_Pct_GDP"),
    ("MS_CentralBank_SovDebt_Share",     "MS_CentralBank_SovDebt_Share"),
    ("MS_Investor_Base_Fragility",       "MS_Investor_Base_Fragility"),
    ("MS_Policy_Backstop",               "MS_Policy_Backstop"),
    ("MS_Reserve_Adequacy",              "MS_Reserve_Adequacy"),
    ("MS_Swap_Line_Access",              "MS_Swap_Line_Access"),
    ("MS_Household_Direct_Equity_e2fe",  "MS_Household_Direct_Equity_Share"),
    ("MS_Insurance_Assets_GDP",          "MS_Insurance_Assets_GDP"),
    ("MS_Pension_Assets_GDP",            "MS_Pension_Assets_GDP"),
    ("MS_US_Holder_Share_Pct",           "MS_US_Holder_Share_Pct"),
    ("MS_Public_Debt_Domestic_Cr_f2ec",  "MS_Public_Debt_Domestic_Creditors_Pct_GDP"),
    ("MS_Public_Debt_Domestic_Cu_3b00",  "MS_Public_Debt_Domestic_Currency_Pct_GDP"),
    ("MS_Public_Debt_External_Cr_fcee",  "MS_Public_Debt_External_Creditors_Pct_GDP"),
    ("MS_Public_Debt_Foreign_Cur_e9e1",  "MS_Public_Debt_Foreign_Currency_Pct_GDP"),
    ("MS_Public_Debt_Foreign_Held_Pct",  "MS_Public_Debt_Foreign_Held_Pct"),
    ("MS_Public_Debt_Local_Curre_d3c7",  "MS_Public_Debt_Local_Currency_Pct"),
    ("MS_Public_Debt_Short_Matur_9fea",  "MS_Public_Debt_Short_Maturity_Pct"),
    ("MS_Public_Debt_Short_Term__896c",  "MS_Public_Debt_Short_Term_Pct_GDP"),
    ("MS_Public_Debt_Total_Pct_GDP",     "MS_Public_Debt_Total_Pct_GDP"),
]

# Redundancy prune: drop highly-correlated, same-concept factors (see
# scripts/econ_drop_list.py). Removing a base sheet also removes its derived
# _CS / _TS / _D12 variants downstream.
_SHEETS_BEFORE = len(SHEETS)
SHEETS = [(s, v) for (s, v) in SHEETS if not should_drop(s)]
logger.info("Econ redundancy prune: %d -> %d sheets (%d redundant dropped).",
            _SHEETS_BEFORE, len(SHEETS), _SHEETS_BEFORE - len(SHEETS))


# ═══════════════════════════════════════════════════════════════════════════════
# Data loading
# ═══════════════════════════════════════════════════════════════════════════════

def load_all_variables(con: duckdb.DuckDBPyConnection) -> dict[str, pd.DataFrame]:
    """
    Load all 146 Econ variables from unified_panel.
    Returns dict: variable_name → wide DataFrame (dates × T2 countries).
    """
    variables = list({var for _, var in SHEETS})
    placeholders = ", ".join(f"'{v}'" for v in variables)

    logger.info("Querying unified_panel for %d variables ...", len(variables))
    raw = con.execute(f"""
        SELECT date, country, value, variable
        FROM unified_panel
        WHERE variable IN ({placeholders})
        ORDER BY variable, date, country
    """).df()

    raw["date"] = pd.to_datetime(raw["date"])
    raw["date"] = raw["date"].dt.to_period("M").dt.to_timestamp()

    # Deduplicate: keep last per (date, country, variable)
    raw = raw.drop_duplicates(subset=["date", "country", "variable"], keep="last")

    wide: dict[str, pd.DataFrame] = {}
    for var, grp in raw.groupby("variable"):
        pivot = grp.pivot_table(index="date", columns="country", values="value", aggfunc="first")
        pivot = pivot.reindex(columns=T2_COUNTRIES)
        pivot = pivot.sort_index()
        wide[var] = pivot

    logger.info("Loaded %d variables, %d missing from DB",
                len(wide), len(variables) - len(wide))
    missing = [v for v in variables if v not in wide]
    if missing:
        logger.warning("Variables not found in unified_panel: %s", missing)

    return wide


# ═══════════════════════════════════════════════════════════════════════════════
# Workbook writing
# ═══════════════════════════════════════════════════════════════════════════════

def write_index_sheet(wb, wide: dict[str, pd.DataFrame]) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet("INDEX", 0)
    headers = ["sheet_id", "sheet_name", "variable", "source",
               "first_date", "last_date", "n_dates", "n_countries_with_data"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    for row_idx, (sheet_name, variable) in enumerate(SHEETS, start=2):
        df = wide.get(variable)
        if df is not None and not df.empty:
            first_date = df.index.min()
            last_date = df.index.max()
            n_dates = len(df.index)
            n_countries = int(df.notna().any().sum())
        else:
            first_date = last_date = None
            n_dates = n_countries = 0

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=sheet_name)
        ws.cell(row=row_idx, column=3, value=variable)
        ws.cell(row=row_idx, column=4, value="unified_panel")
        ws.cell(row=row_idx, column=5, value=first_date)
        ws.cell(row=row_idx, column=6, value=last_date)
        ws.cell(row=row_idx, column=7, value=n_dates)
        ws.cell(row=row_idx, column=8, value=n_countries)

    for col_letter, width in zip("ABCDEFGH", [8, 33, 40, 16, 12, 12, 8, 22]):
        ws.column_dimensions[col_letter].width = width


def write_data_sheet(wb, sheet_name: str, df: pd.DataFrame | None) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 10.33

    # Header row
    ws.cell(row=1, column=1, value="date")
    for col_idx, country in enumerate(T2_COUNTRIES, start=2):
        ws.cell(row=1, column=col_idx, value=country)

    if df is None or df.empty:
        return

    # Data rows
    for row_idx, dt in enumerate(df.index, start=2):
        date_cell = ws.cell(row=row_idx, column=1, value=dt.to_pydatetime())
        date_cell.number_format = "yyyy-mm-dd"
        for col_idx, country in enumerate(T2_COUNTRIES, start=2):
            value = df.at[dt, country] if country in df.columns else None
            ws.cell(row=row_idx, column=col_idx,
                    value=None if (value is None or (isinstance(value, float) and pd.isna(value))) else value)


# ═══════════════════════════════════════════════════════════════════════════════
# Backup + parquet snapshot
# ═══════════════════════════════════════════════════════════════════════════════

def backup_local_parquet() -> None:
    if not LOCAL_PARQUET.exists():
        return
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    backup_dir = BACKUPS_DIR / timestamp
    backup_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(LOCAL_PARQUET, backup_dir / LOCAL_PARQUET.name)
    logger.info("Backed up existing parquet to %s", backup_dir)


def save_local_parquet(wide: dict[str, pd.DataFrame]) -> None:
    frames = []
    for var, df in wide.items():
        melted = df.reset_index().melt(id_vars="date", var_name="country", value_name="value")
        melted["variable"] = var
        frames.append(melted)
    if not frames:
        return
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.dropna(subset=["value"])
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    combined.to_parquet(LOCAL_PARQUET, index=False)
    logger.info("Saved local parquet: %s (%d rows)", LOCAL_PARQUET, len(combined))


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="Preview only — no files written.")
    ap.add_argument("--check", action="store_true",
                    help="Report on existing output files without rebuilding.")
    ap.add_argument("--force", action="store_true",
                    help="Rebuild even if Econ.xlsx is newer than the DuckDB.")
    args = ap.parse_args()

    # ── Up-to-date guard ──────────────────────────────────────────────────
    # Output is Econ.xlsx; data source is asado.duckdb. If xlsx mtime ≥ duckdb
    # mtime, no upstream change has landed and we can skip the entire run.
    if (
        not args.force
        and not args.dry_run
        and not args.check
        and OUTPUT_XLSX.exists()
        and DUCKDB_PATH.exists()
        and OUTPUT_XLSX.stat().st_mtime >= DUCKDB_PATH.stat().st_mtime
    ):
        logger.info(
            "Econ.xlsx is newer than asado.duckdb — no upstream change. "
            "Skipping. Use --force to rebuild."
        )
        return 0

    if args.check:
        if LOCAL_PARQUET.exists():
            df = pd.read_parquet(LOCAL_PARQUET)
            logger.info("Local parquet: %d rows, dates %s → %s, %d variables",
                        len(df),
                        str(df["date"].min())[:10], str(df["date"].max())[:10],
                        df["variable"].nunique())
        else:
            logger.warning("Local parquet not found: %s", LOCAL_PARQUET)
        if OUTPUT_XLSX.exists():
            stat = OUTPUT_XLSX.stat()
            logger.info("Econ.xlsx: %.1f MB, modified %s",
                        stat.st_size / 1_048_576,
                        datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"))
        else:
            logger.warning("Econ.xlsx not found: %s", OUTPUT_XLSX)
        return 0

    if not DUCKDB_PATH.exists():
        logger.error("DuckDB not found: %s", DUCKDB_PATH)
        return 1

    # ── Load ─────────────────────────────────────────────────────────────────
    with duckdb.connect(str(DUCKDB_PATH), read_only=True) as con:
        wide = load_all_variables(con)

    if args.dry_run:
        total_rows = sum(len(df) * len(T2_COUNTRIES) for df in wide.values())
        logger.info("[dry-run] Would write %d sheets to %s", len(SHEETS) + 1, OUTPUT_XLSX)
        logger.info("[dry-run] Approx %d data cells across %d variables", total_rows, len(wide))
        return 0

    # ── Backup + local parquet ───────────────────────────────────────────────
    backup_local_parquet()
    save_local_parquet(wide)

    # ── Build workbook ───────────────────────────────────────────────────────
    from openpyxl import Workbook

    logger.info("Building workbook: INDEX + %d data sheets ...", len(SHEETS))
    wb = Workbook()
    wb.remove(wb.active)

    write_index_sheet(wb, wide)

    missing_count = 0
    for sheet_name, variable in SHEETS:
        df = wide.get(variable)
        if df is None:
            logger.warning("No data for variable '%s' — writing empty sheet '%s'",
                           variable, sheet_name)
            missing_count += 1
        write_data_sheet(wb, sheet_name, df)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    size_mb = OUTPUT_XLSX.stat().st_size / 1_048_576
    logger.info(
        "Wrote %s (%.1f MB, %d sheets: 1 INDEX + %d data, %d empty)",
        OUTPUT_XLSX, size_mb, len(wb.sheetnames), len(SHEETS), missing_count,
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
