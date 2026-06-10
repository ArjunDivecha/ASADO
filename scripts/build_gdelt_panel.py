#!/usr/bin/env python3
"""
=============================================================================
SCRIPT NAME: build_gdelt_panel.py
=============================================================================

DESCRIPTION:
    Reads the GDELT Deep monthly country-level feature panel parquet, validates
    it (deduplicates, drops incomplete current month), backs up any existing
    local parquet snapshot, and writes a fresh local parquet copy. Then builds
    an Excel workbook (GDELT.xlsx) containing:
      - A README sheet documenting the architecture and variable families
      - A variable dictionary sheet with family/stage/definition for every column
      - One wide-format data sheet per variable (dates as rows, countries as
        columns), covering core tone signals plus a curated subset of deep
        features (theme shares/deltas, event aggregates) pruned by the GDELT
        keep-list rule (gdelt_keep_list.py).
    Supports --dry-run (preview, no writes), --check (report on existing output),
    and --force (rebuild even if output is up-to-date).

INPUT FILES:
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/GDELT/Deep/data/features/country_signal_monthly_deep.parquet
        Upstream monthly country-level GDELT Deep panel produced by
        build_monthly_metronome.py. The primary data source for the workbook.
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/Data/processed/gdelt_workbook_panel.parquet
        Existing local parquet snapshot read during --check mode for status
        reporting. Backed up before overwrite in normal mode.

OUTPUT FILES:
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/Data/processed/gdelt_workbook_panel.parquet
        Local parquet snapshot of the validated panel. Backed up before
        overwrite (see backup path below).
    /Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT/GDELT.xlsx
        Wide-format Excel workbook with two README documentation sheets
        followed by one wide-format data sheet per variable (dates x countries).
        Country order in column headers matches the T2 layout. Used by the
        T2 GDELT downstream pipeline.
    /Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/Data/backups/{timestamp}/gdelt_workbook_panel.parquet
        Backup copy of the previous local parquet, created before overwriting
        the current one.

VERSION: 1.0
LAST UPDATED: 2026-06-05
AUTHOR: Arjun Divecha

DEPENDENCIES:
    - pandas
    - openpyxl
    - pyarrow

USAGE:
    python build_gdelt_panel.py                # normal run
    python build_gdelt_panel.py --dry-run      # preview, no writes
    python build_gdelt_panel.py --check        # report on existing output
    python build_gdelt_panel.py --force        # rebuild regardless of mtime

NOTES:
    - Upstream parquet path is hard-coded to the canonical GDELT Deep location.
      If the file is missing, the script fails loudly (no fallback).
    - Output GDELT.xlsx path is hard-coded to the T2 GDELT directory.
    - Only fully-completed months are exported (partial current month excluded).
    - Deep features (theme_*, gcam_*, event_*) are pruned by
      gdelt_keep_list.py's keep_deep_feature() rule before writing to the
      workbook.
    - An mtime guard compares GDELT.xlsx mtime against the upstream parquet
      mtime to skip runs when no upstream change has landed.
=============================================================================
"""

from __future__ import annotations

import argparse
import logging
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import pandas as pd

# Canonical GDELT Deep keep-list (single source of truth for the prune rule).
# Resolves whether build_gdelt_panel runs as a script (sys.path[0]=scripts/) or
# is imported with the repo root on the path.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from gdelt_keep_list import canonical_name, keep_deep_feature  # noqa: E402

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"
PROCESSED_DIR = DATA_DIR / "processed"
BACKUPS_DIR = DATA_DIR / "backups"

UPSTREAM_PARQUET = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Working/GDELT/Deep"
    "/data/features/country_signal_monthly_deep.parquet"
)
OUTPUT_XLSX = Path(
    "/Users/arjundivecha/Dropbox/AAA Backup/A Complete/T2 GDELT/GDELT.xlsx"
)
LOCAL_PARQUET = PROCESSED_DIR / "gdelt_workbook_panel.parquet"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ── Country bucket → ISO3 mapping (T2 order) ────────────────────────────────
COUNTRY_BUCKETS = [
    ("Singapore", "SGP"),     ("Australia", "AUS"),      ("Canada", "CAN"),
    ("Germany", "DEU"),       ("Japan", "JPN"),          ("Switzerland", "CHE"),
    ("U.K.", "GBR"),          ("U.S. NASDAQ", "USA"),    ("U.S.", "USA"),
    ("France", "FRA"),        ("Netherlands", "NLD"),    ("Sweden", "SWE"),
    ("Italy", "ITA"),         ("China A", "CHN"),        ("Chile", "CHL"),
    ("Indonesia", "IDN"),     ("Philippines", "PHL"),    ("Poland", "POL"),
    ("US SmallCap", "USA"),   ("Malaysia", "MYS"),       ("Taiwan", "TWN"),
    ("Mexico", "MEX"),        ("Korea", "KOR"),          ("Brazil", "BRA"),
    ("South Africa", "ZAF"),  ("Denmark", "DNK"),        ("India", "IND"),
    ("China H", "CHN"),       ("Hong Kong", "HKG"),      ("Thailand", "THA"),
    ("Turkey", "TUR"),        ("Spain", "ESP"),          ("Vietnam", "VNM"),
    ("Saudi Arabia", "SAU"),
]

# ── Core signal columns (preserved from original workbook order) ─────────────
CORE_INDICATORS = [
    "monthly_metronome", "monthly_risk", "monthly_defensive",
    "monthly_metronome_rank_pct", "monthly_risk_rank_pct", "monthly_defensive_rank_pct",
    "sentiment_fast", "sentiment_slow", "sentiment_trend",
    "attention_fast", "attention_slow", "attention_trend",
    "risk_fast", "dispersion_fast",
    "local_tone_fast", "foreign_tone_fast", "local_foreign_gap",
    "sentiment_fast_z", "sentiment_slow_z", "sentiment_trend_z",
    "attention_fast_z", "attention_slow_z", "attention_trend_z",
    "risk_fast_z", "dispersion_fast_z",
    "local_tone_fast_z", "foreign_tone_fast_z", "local_foreign_gap_z",
    "country_news_sentiment", "country_news_risk",
    "country_news_sentiment_raw", "country_news_risk_raw",
    "country_news_attention", "local_attention_share",
    "country_news_sentiment_x_attention",
    "local_tone", "foreign_tone", "attention_shock",
    "tone_dispersion", "tone_wavg_wordcount",
    "tone_mean", "tone_p50", "positive_mean", "negative_mean",
    "polarity_mean", "n_articles",
]

# ── Metadata columns excluded from data sheets ───────────────────────────────
SKIP_PREFIXES = (
    "month_obs_", "month_calendar_", "month_day_status_",
    "month_gkg_", "signal_month",
)
SKIP_COLS = {
    "date", "country_iso3", "country_name", "country_code_gdelt",
    "row_date", "month_end_date_used", "signal_month_end_date",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Family / stage classifiers
# ═══════════════════════════════════════════════════════════════════════════════

def classify_family(col: str) -> str:
    if col.startswith("theme_"):  return "THEME ATTENTION"
    if col.startswith("gcam_"):   return "GCAM EMOTIONAL DIMENSIONS"
    if col.startswith("event_"):  return "EVENT AGGREGATES"
    return "CORE TONE SIGNALS"


def classify_stage(col: str) -> str:
    if col.endswith("_z") or col.endswith("_rank_pct"):
        return "monthly → standardized"
    if any(col.endswith(s) for s in ["_fast", "_slow", "_trend"]):
        return "monthly → EWMA"
    if "_delta" in col:
        return "monthly → differenced"
    if col.startswith("theme_") or col.startswith("gcam_") or col.startswith("event_"):
        return "article-level → country-day → monthly snapshot"
    return "daily → monthly snapshot"


# ═══════════════════════════════════════════════════════════════════════════════
# Variable definitions
# ═══════════════════════════════════════════════════════════════════════════════

def build_variable_definitions(frame: pd.DataFrame) -> dict:
    defs: dict[str, str] = {}

    core_defs = {
        "monthly_metronome": "Primary monthly composite. 0.35*sentiment_fast_z + 0.20*sentiment_slow_z + 0.20*sentiment_trend_z + 0.15*attention_fast_z − 0.10*risk_fast_z",
        "monthly_risk": "Risk composite. 0.45*risk_fast_z + 0.30*dispersion_fast_z − 0.15*sentiment_fast_z − 0.10*foreign_tone_fast_z",
        "monthly_defensive": "Defensive score = −1 * monthly_risk",
        "sentiment_fast": "EWMA(span=5) of country_news_sentiment_raw",
        "sentiment_slow": "EWMA(span=20) of country_news_sentiment_raw",
        "sentiment_trend": "sentiment_fast − sentiment_slow",
        "attention_fast": "EWMA(span=5) of local_attention_share",
        "attention_slow": "EWMA(span=20) of local_attention_share",
        "attention_trend": "attention_fast − attention_slow",
        "risk_fast": "EWMA(span=10) of country_news_risk_raw",
        "dispersion_fast": "EWMA(span=10) of tone_dispersion",
        "local_tone_fast": "EWMA(span=5) of local_tone",
        "foreign_tone_fast": "EWMA(span=5) of foreign_tone",
        "local_foreign_gap": "local_tone_fast − foreign_tone_fast",
        "country_news_sentiment": "Daily trailing z-score of country_news_sentiment_raw, month-end snapshot",
        "country_news_risk": "Daily trailing z-score of country_news_risk_raw, month-end snapshot",
        "country_news_sentiment_raw": "Base raw sentiment = local_tone (fallback tone_wavg_wordcount)",
        "country_news_risk_raw": "Base raw risk = −sentiment_raw + 0.5*tone_dispersion",
        "country_news_attention": "Log-scaled article attention = log(1 + local_n_articles)",
        "local_attention_share": "local_n_articles / local_source_total_articles",
        "local_tone": "Average tone of local-source articles",
        "foreign_tone": "Average tone of foreign-source articles",
        "attention_shock": "Daily trailing z-score of country_news_attention",
        "tone_dispersion": "Spread of tone values across articles",
        "tone_wavg_wordcount": "Word-count-weighted average tone",
        "tone_mean": "Simple average tone across all articles",
        "tone_p50": "Median tone (50th percentile)",
        "positive_mean": "Average positive-tone component",
        "negative_mean": "Average negative-tone component",
        "polarity_mean": "Average polarity component",
        "n_articles": "Total article count for the country-day snapshot",
    }
    defs.update(core_defs)

    for base in ["sentiment", "attention", "risk", "dispersion",
                 "local_tone", "foreign_tone", "local_foreign_gap"]:
        for suffix in ["fast", "slow", "trend", ""]:
            name = f"{base}_{suffix}_z" if suffix else f"{base}_z"
            if name in defs:
                continue
            base_name = f"{base}_{suffix}" if suffix else base
            defs[name] = f"Trailing within-country z-score of {base_name} (window=24mo, min=6mo, shift=1)"

    defs["local_foreign_gap_z"] = "Trailing z-score of local_foreign_gap (24mo window)"
    defs["lf_gap_z"] = defs["local_foreign_gap_z"]

    for name in ["monthly_metronome_rank_pct", "monthly_risk_rank_pct", "monthly_defensive_rank_pct"]:
        base = name.replace("_rank_pct", "")
        defs[name] = f"Cross-sectional percentile rank of {base} within each month (0 to 1)"

    for col in [c for c in frame.columns if c.startswith("theme_")]:
        theme_name = col.replace("theme_", "").replace("_share", "").replace("_delta", "")
        if col.endswith("_share_delta"):
            defs[col] = f"Month-over-month change in theme_share for '{theme_name}'"
        elif col.endswith("_share"):
            defs[col] = f"Share of articles mentioning theme '{theme_name}' in this country-month"
        else:
            defs[col] = f"Theme '{theme_name}' indicator"

    for col in [c for c in frame.columns if c.startswith("gcam_")]:
        dim = col.replace("gcam_", "")
        if dim.endswith("_fast"):
            defs[col] = f"EWMA(span=5) of gcam_{dim[:-5]} — fast GCAM signal"
        elif dim.endswith("_slow"):
            defs[col] = f"EWMA(span=20) of gcam_{dim[:-5]} — slow GCAM signal"
        elif dim.endswith("_trend"):
            defs[col] = f"gcam_{dim[:-6]}_fast − gcam_{dim[:-6]}_slow — GCAM momentum"
        elif dim.endswith("_z"):
            defs[col] = f"Trailing z-score(24mo) of gcam_{dim[:-2]}"
        else:
            defs[col] = f"Country-month mean of GCAM dimension '{dim}'"

    for col in [c for c in frame.columns if c.startswith("event_")]:
        feat = col.replace("event_", "")
        if feat.endswith("_fast"):
            defs[col] = f"EWMA(span=5) of event_{feat[:-5]} — fast event signal"
        elif feat.endswith("_trend"):
            defs[col] = f"event_{feat[:-6]}_fast − lagged — event momentum"
        elif feat.endswith("_z"):
            defs[col] = f"Trailing z-score(24mo) of event_{feat[:-2]}"
        elif "quad" in feat:
            defs[col] = f"Count of QuadClass={feat[-1]} events (1=VerbalCoop, 2=MaterialCoop, 3=VerbalConflict, 4=MaterialConflict)"
        elif "goldstein" in feat:
            defs[col] = f"Goldstein cooperation/conflict score (−10 to +10): {feat}"
        elif "persistence" in feat:
            defs[col] = f"Count of mentions today of events from {feat.replace('persistence_','').replace('d','')} days ago"
        elif "root_" in feat:
            defs[col] = f"Count of events with EventRootCode '{feat.replace('root_','').replace('_n','')}'"
        else:
            defs[col] = f"Event aggregate: {feat}"

    return defs


# ═══════════════════════════════════════════════════════════════════════════════
# Workbook writing helpers
# ═══════════════════════════════════════════════════════════════════════════════

README_TEXT = """
GDELT COUNTRY NEWS MODALITY — MONTHLY FEATURE PANEL
=====================================================

ARCHITECTURE
This workbook contains monthly country-level features built from the GDELT
Global Knowledge Graph (GKG v2). Features are organized into four families:

FAMILY 1: CORE TONE SIGNALS (~45 variables)
  Derived from GKG V2Tone field — the 8 base tone measurements:
  tone_mean, positive_mean, negative_mean, polarity_mean, local_tone,
  foreign_tone, tone_dispersion, n_articles.
  → Fast/slow/trend EWMA decompositions (spans 5 / 20 months)
  → Trailing within-country z-scores (24-month window, min 6 months)
  → Monthly composites: metronome, risk, defensive
  → Cross-sectional rank percentiles within each month

FAMILY 2: THEME ATTENTION (~568 variables)
  Source: GKG V2Themes — 284 pre-curated themes from the GDELT THEMES taxonomy.
  For each country-month:
    theme_<NAME>_share  = fraction of articles mentioning the theme
    theme_<NAME>_share_delta = month-over-month change in share
  The share measures topic attention. The delta captures attention SHIFTS.

FAMILY 3: GCAM EMOTIONAL DIMENSIONS (~300 variables)
  Source: GKG GCAM field — 40 curated sentiment dictionaries applied to
  every article, aggregated to country-day means, then to monthly snapshots.
  Four sub-blocks: Loughran-McDonald (6 dims), WordNet-Affect 1.0 (11 dims),
  Harvard IV-4 General Inquirer (18 dims), Value-based (40 dims).

FAMILY 4: EVENT AGGREGATES (~72 variables)
  Source: GDELT CAMEO EVENTS + EVENTMENTIONS tables.
  Country attribution via ActionGeo_CountryCode (FIPS→ISO3 mapped).

NAMING CONVENTIONS
  theme_<NAME>_share[_delta]  — Theme attention share and delta
  gcam_<DIM>_<treatment>      — GCAM dimension (fast/slow/trend/z)
  event_<FEAT>_<treatment>    — Event feature (fast/trend/z)

FREQUENCY
  All data is calendar-monthly. Value for month M = last calendar day of M.

NO LOOK-AHEAD
  Every feature for month M is computable from data published on or before
  the last day of month M. Z-scores use trailing windows shifted by 1 month.
"""


def write_readme(wb, frame: pd.DataFrame) -> None:
    from openpyxl import Workbook  # noqa: F401 — wb already created
    ws = wb.create_sheet("README", 0)
    for row_idx, line in enumerate(README_TEXT.strip().split("\n"), start=1):
        ws.cell(row=row_idx, column=1, value=line)

    gap = len(README_TEXT.strip().split("\n")) + 2
    ws.cell(row=gap, column=1, value="COVERAGE SUMMARY")
    min_date = frame["date"].min()
    max_date = frame["date"].max()
    min_str = min_date.strftime("%Y-%m-%d") if hasattr(min_date, "strftime") else str(min_date)[:10]
    max_str = max_date.strftime("%Y-%m-%d") if hasattr(max_date, "strftime") else str(max_date)[:10]

    families: dict[str, int] = {}
    for c in frame.columns:
        fam = classify_family(c)
        families[fam] = families.get(fam, 0) + 1

    metrics = [
        f"Date range: {min_str} to {max_str}",
        f"Countries: {len(COUNTRY_BUCKETS)} buckets",
    ] + [f"{fam}: {cnt} variables" for fam, cnt in sorted(families.items())] + [
        f"Total variables: {sum(families.values())}"
    ]

    for offset, metric in enumerate(metrics, start=1):
        ws.cell(row=gap + offset, column=1, value=f"  • {metric}")

    ws.column_dimensions["A"].width = 100


def write_variable_dictionary(wb, frame: pd.DataFrame, definitions: dict) -> None:
    from openpyxl.styles import Font
    ws = wb.create_sheet("README_VARIABLES", 1)
    headers = ["sheet_name", "pipeline_column", "family", "stage", "definition"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    fam_order = {"CORE TONE SIGNALS": 0, "THEME ATTENTION": 1,
                 "GCAM EMOTIONAL DIMENSIONS": 2, "EVENT AGGREGATES": 3}
    # Only document the columns actually written as data sheets: core indicators
    # plus the kept deep features (post-prune). Keeps the dictionary honest.
    core_set = set(CORE_INDICATORS)
    written = [c for c in frame.columns if c in core_set or keep_deep_feature(c)]
    ordered = sorted(
        [(fam_order.get(classify_family(c), 99), c) for c in written]
    )

    row = 2
    for _, col in ordered:
        sheet = col[:31] if col in core_set else canonical_name(col)
        ws.cell(row=row, column=1, value=sheet)
        ws.cell(row=row, column=2, value=col)
        ws.cell(row=row, column=3, value=classify_family(col))
        ws.cell(row=row, column=4, value=classify_stage(col))
        ws.cell(row=row, column=5, value=definitions.get(col, f"See {classify_family(col)} family documentation in README"))
        row += 1

    for col_idx, width in enumerate([32, 40, 28, 34, 100], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width


def write_data_sheet(wb, sheet_name: str, frame: pd.DataFrame, col: str) -> None:
    ws = wb.create_sheet(sheet_name)
    panel = frame.pivot_table(
        index="row_date", columns="country_iso3", values=col, aggfunc="first"
    ).sort_index()

    needed = sorted({iso3 for _, iso3 in COUNTRY_BUCKETS})
    panel = panel.reindex(columns=needed)

    ws.column_dimensions["A"].width = 10.33

    ws.cell(row=1, column=1, value=None)
    for idx, (label, _) in enumerate(COUNTRY_BUCKETS, start=2):
        ws.cell(row=1, column=idx, value=label)

    num_fmt = "0" if "n_articles" in col else "0.000000"
    for row_idx, dt in enumerate(panel.index, start=2):
        date_cell = ws.cell(row=row_idx, column=1, value=dt.to_pydatetime())
        date_cell.number_format = "mm-dd-yy"
        for col_idx, (_, iso3) in enumerate(COUNTRY_BUCKETS, start=2):
            value = panel.at[dt, iso3] if iso3 in panel.columns else None
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=None if (value is None or pd.isna(value)) else value)
            cell.number_format = num_fmt


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def load_and_validate() -> pd.DataFrame:
    """Load upstream parquet and return a clean, deduplicated frame."""
    if not UPSTREAM_PARQUET.exists():
        logger.error("Upstream parquet not found: %s", UPSTREAM_PARQUET)
        raise FileNotFoundError(UPSTREAM_PARQUET)

    logger.info("Loading %s", UPSTREAM_PARQUET)
    frame = pd.read_parquet(UPSTREAM_PARQUET)
    frame["date"] = pd.to_datetime(frame["date"])

    if "signal_month_end_date" in frame.columns:
        frame["row_date"] = pd.to_datetime(frame["signal_month_end_date"])
    elif "signal_month" in frame.columns:
        frame["row_date"] = pd.to_datetime(
            frame["signal_month"].astype(str).str[:7] + "-01"
        )
    else:
        frame["row_date"] = frame["date"]

    frame = frame.dropna(subset=["country_iso3"])
    frame["country_iso3"] = frame["country_iso3"].astype(str).str.strip()
    frame = frame[frame["country_iso3"] != ""]

    # Drop incomplete current month
    max_obs = frame["date"].max()
    max_month_end = max_obs.to_period("M").to_timestamp(how="end").normalize()
    if max_obs < max_month_end:
        completed = (max_obs.to_period("M") - 1).to_timestamp(how="end").normalize()
        frame = frame[frame["row_date"] <= completed]

    # Deduplicate: keep last row per (row_date, country_iso3)
    frame = frame.sort_values(["row_date", "country_iso3", "date"])
    frame = frame.drop_duplicates(subset=["row_date", "country_iso3"], keep="last")

    logger.info(
        "Validated panel: %d dates, %d countries, %d rows",
        frame["row_date"].nunique(),
        frame["country_iso3"].nunique(),
        len(frame),
    )
    return frame


def backup_local_parquet() -> None:
    if not LOCAL_PARQUET.exists():
        return
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    backup_dir = BACKUPS_DIR / timestamp
    backup_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(LOCAL_PARQUET, backup_dir / LOCAL_PARQUET.name)
    logger.info("Backed up existing parquet to %s", backup_dir)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="Preview only — no files written.")
    ap.add_argument("--check", action="store_true",
                    help="Report on existing output files without rebuilding.")
    ap.add_argument("--force", action="store_true",
                    help="Rebuild even if GDELT.xlsx is newer than the upstream parquet.")
    args = ap.parse_args()

    # ── Up-to-date guard ──────────────────────────────────────────────────
    # Output GDELT.xlsx is built from UPSTREAM_PARQUET. If xlsx mtime ≥ parquet
    # mtime, no upstream change has landed and we can skip the entire run.
    if (
        not args.force
        and not args.dry_run
        and not args.check
        and OUTPUT_XLSX.exists()
        and UPSTREAM_PARQUET.exists()
        and OUTPUT_XLSX.stat().st_mtime >= UPSTREAM_PARQUET.stat().st_mtime
    ):
        logger.info(
            "GDELT.xlsx is newer than %s — no upstream change. "
            "Skipping. Use --force to rebuild.", UPSTREAM_PARQUET.name
        )
        return 0

    if args.check:
        if LOCAL_PARQUET.exists():
            df = pd.read_parquet(LOCAL_PARQUET)
            logger.info("Local parquet: %d rows, %d cols, dates %s → %s",
                        len(df), len(df.columns),
                        str(df["date"].min())[:10], str(df["date"].max())[:10])
        else:
            logger.warning("Local parquet not found: %s", LOCAL_PARQUET)

        if OUTPUT_XLSX.exists():
            stat = OUTPUT_XLSX.stat()
            logger.info("GDELT.xlsx: %.1f MB, modified %s",
                        stat.st_size / 1_048_576,
                        datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"))
        else:
            logger.warning("GDELT.xlsx not found: %s", OUTPUT_XLSX)
        return 0

    # ── Load & validate ──────────────────────────────────────────────────────
    try:
        frame = load_and_validate()
    except FileNotFoundError:
        return 1

    if args.dry_run:
        logger.info("[dry-run] Would write %d rows to %s", len(frame), OUTPUT_XLSX)
        logger.info("[dry-run] Would write local parquet to %s", LOCAL_PARQUET)
        return 0

    # ── Backup + save local parquet ──────────────────────────────────────────
    backup_local_parquet()
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    frame.to_parquet(LOCAL_PARQUET, index=False)
    logger.info("Saved local parquet: %s", LOCAL_PARQUET)

    # ── Build workbook ───────────────────────────────────────────────────────
    from openpyxl import Workbook

    # Variable columns (exclude metadata)
    var_cols = [
        c for c in frame.columns
        if c not in SKIP_COLS and not any(c.startswith(p) for p in SKIP_PREFIXES)
    ]
    core_set = set(CORE_INDICATORS)
    available_core = [c for c in CORE_INDICATORS if c in var_cols]
    # GDELT prune: of the deep (theme_/gcam_/event_) feature columns, keep only
    # the curated keep-list — ~24 interpretable themes ({_share,_share_delta})
    # and ~16 CAMEO/Goldstein event aggregates. All gcam_* and the long theme
    # tail are dropped. See scripts/gdelt_keep_list.py for the rule + rationale.
    all_deep = [c for c in var_cols if c not in core_set]
    deep_cols = [c for c in all_deep if keep_deep_feature(c)]
    logger.info(
        "GDELT prune: %d deep feature cols -> %d kept (%d themes, %d events); "
        "%d dropped (gcam + theme tail).",
        len(all_deep), len(deep_cols),
        sum(c.startswith("theme_") for c in deep_cols),
        sum(c.startswith("event_") for c in deep_cols),
        len(all_deep) - len(deep_cols),
    )

    logger.info(
        "Building workbook: %d core + %d deep = %d data sheets",
        len(available_core), len(deep_cols), len(available_core) + len(deep_cols),
    )

    definitions = build_variable_definitions(frame)
    wb = Workbook()
    wb.remove(wb.active)

    write_readme(wb, frame)
    write_variable_dictionary(wb, frame, definitions)

    for col in available_core:
        write_data_sheet(wb, col[:31], frame, col)

    # Deep features use canonical_name(): clean, unique, <=31-char sheet names.
    # The sheet name becomes the downstream variable base (<sheet>_CS/_TS), so
    # this also repairs the legacy 31-char truncation collisions (e.g. the old
    # theme_TAX_MILITARY_TITLE_OFFICE1/2/3 garbage).
    for col in sorted(deep_cols):
        write_data_sheet(wb, canonical_name(col), frame, col)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    logger.info(
        "Wrote %s (%.1f MB, %d sheets: 2 docs + %d data)",
        OUTPUT_XLSX,
        OUTPUT_XLSX.stat().st_size / 1_048_576,
        len(wb.sheetnames),
        len(available_core) + len(deep_cols),
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
