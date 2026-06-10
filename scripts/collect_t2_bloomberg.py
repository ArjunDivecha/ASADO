"""
=============================================================================
SCRIPT NAME: collect_t2_bloomberg.py
=============================================================================

DESCRIPTION:
    Generates the T2 Bloomberg master workbook DIRECTLY from the live Bloomberg
    API (blpapi via OpusBloomberg), replacing the hand-maintained Excel dump
    "Country Bloomberg Data Master T.xlsx". Every data sheet in that dump is a
    uniform BDH pull (=BDH(ticker, field, "1999-12-31", "", FX=USD, Per=M)),
    and the 3 weight sheets are pure calculations — so ASADO can reproduce the
    whole workbook with no human and no janky spreadsheet in the middle.

    Recipe is frozen in scripts/config/t2_bbg_manifest.json (extracted once from
    the dump): per-sheet (ticker, field) lists + the MCAP split factors.

    Reproduction, validated to the decimal vs the dump (e.g. MXAU Index PX_LAST
    USD 2026-05-29 = 1232.97289 both live and dump):
      - 38 direct sheets: bbg.hist(ticker, field, start, today, MONTHLY,
        currency="USD"), forward-filled (Fill=P), assembled date x ticker.
      - MCAP Adj   = MCAP * split_factor   (splits US into 3, China into 2)
      - Mcap Weights = MCAP Adj / rowsum(MCAP Adj)
      - Master     = static ETF->index map + split factors (sheet 1; build_t2_master skips it)

INPUT FILES:
    - scripts/config/t2_bbg_manifest.json   (the frozen recipe)
    - Bloomberg live API (OpusBloomberg) — Terminal must be up on Parallels

OUTPUT FILES:
    - --out  (default: Data/work/t2/T2 Bloomberg Master.xlsx)
      Same sheet structure as the legacy dump; build_t2_master reads this.

VERSION: 1.0
LAST UPDATED: 2026-06-07
AUTHOR: Arjun Divecha (built by Claude Code)

DEPENDENCIES:
    - OpusBloomberg (bbg.BBG, bloomberg_setup) — run in the OpusBloomberg conda env
    - pandas, numpy, openpyxl

USAGE (per bloomberg-skill — OpusBloomberg conda env, NOT the ASADO venv):
    conda run -p "/Users/arjundivecha/Dropbox/AAA Backup/A Working/OpusBloomberg/.venv" \
      python "/Users/arjundivecha/Dropbox/AAA Backup/A Working/ASADO/scripts/collect_t2_bloomberg.py"
    # options:
    #   --sheets "PX_LAST,MCAP"   build only some sheets (testing)
    #   --out PATH                output workbook path
    #   --start YYYYMMDD          history start (default from manifest, 19991231)

NOTES:
    - ~1,292 BDH series pulls (deduped); each full-history series = 1 cap hit.
    - Follows the bloomberg-skill flow: bloomberg_setup() (retry once), with BBG().
=============================================================================
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, '/Users/arjundivecha/Dropbox/AAA Backup/A Working/OpusBloomberg')
import numpy as np
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent.parent
MANIFEST = BASE_DIR / "scripts" / "config" / "t2_bbg_manifest.json"
MANIFEST_DAILY = BASE_DIR / "scripts" / "config" / "t2_bbg_manifest_daily.json"
DEFAULT_OUT = BASE_DIR / "Data" / "work" / "t2" / "T2 Bloomberg Master.xlsx"
DEFAULT_OUT_DAILY = BASE_DIR / "Data" / "work" / "t2" / "T2 Bloomberg Master Daily.xlsx"
CALC_SHEETS = ["MCAP Adj", "Mcap Weights"]   # computed; Master is written as sheet 1


def _setup_with_retry():
    from bbg import bloomberg_setup
    try:
        bloomberg_setup()
    except Exception as e:
        print(f"  bloomberg_setup attempt 1 failed ({e}); retrying once...", flush=True)
        bloomberg_setup()


def rows_to_monthly(rows, field):
    """BDH rows -> pd.Series indexed by monthly PERIOD (one obs per calendar
    month — collapses per-ticker month-end day variation, e.g. 05-29 vs 05-31)."""
    if not rows:
        return pd.Series(dtype=float)
    idx = pd.to_datetime([r["date"] for r in rows])
    vals = pd.to_numeric([r.get(field) for r in rows], errors="coerce")
    s = pd.Series(vals, index=idx).sort_index()
    s.index = s.index.to_period("M")
    return s[~s.index.duplicated(keep="last")]   # one value per month


def rows_to_daily(rows, field):
    """BDH rows -> pd.Series indexed by DAILY date (one obs per trading day).
    Daily analog of rows_to_monthly: NO period collapse — every trading day kept."""
    if not rows:
        return pd.Series(dtype=float)
    idx = pd.to_datetime([r["date"] for r in rows])
    vals = pd.to_numeric([r.get(field) for r in rows], errors="coerce")
    s = pd.Series(vals, index=idx).sort_index()
    return s[~s.index.duplicated(keep="last")]   # one value per day


def main() -> int:
    ap = argparse.ArgumentParser(description="Build the T2 Bloomberg master workbook from live blpapi.")
    ap.add_argument("--out", default="", help="Output workbook path (default depends on --daily).")
    ap.add_argument("--daily", action="store_true",
                    help="Daily mode: use the daily manifest, pull DAILY periodicity, keep every trading day.")
    ap.add_argument("--manifest", default="", help="Explicit manifest JSON path (overrides --daily selection).")
    ap.add_argument("--sheets", default="", help="Comma-separated subset of direct sheets (testing).")
    ap.add_argument("--start", default="", help="Override history start (YYYYMMDD).")
    ap.add_argument("--end", default="", help="Override history end (YYYYMMDD); default today.")
    args = ap.parse_args()

    manifest_path = Path(args.manifest) if args.manifest else (MANIFEST_DAILY if args.daily else MANIFEST)
    manifest = json.loads(manifest_path.read_text())
    bp = manifest["bdh_params"]
    freq = bp.get("periodicity", "MONTHLY").upper()
    is_daily = freq == "DAILY"
    out_path = Path(args.out) if args.out else (DEFAULT_OUT_DAILY if is_daily else DEFAULT_OUT)
    start = args.start or bp["start"]
    end = args.end or date.today().strftime("%Y%m%d")
    currency = bp["currency"]
    only = {s.strip() for s in args.sheets.split(",") if s.strip()}

    sheets = manifest["sheets"]
    if only:
        sheets = [s for s in sheets if s["sheet"] in only]

    # group unique tickers BY FIELD -> one batched HistoricalDataRequest per field
    # (BLPAPI accepts many securities per request and splits internally; this turns
    #  ~1,292 single-security round-trips into ~30, the way Excel's BDH batches).
    from collections import defaultdict
    import time as _time
    by_field = defaultdict(set)
    for s in sheets:
        for c in s["columns"]:
            by_field[c["field"]].add(c["ticker"])
    npairs = sum(len(v) for v in by_field.values())
    print(f"== {npairs} ticker-field series in {len(by_field)} batched requests (by field), "
          f"{start}->{end}, {currency}, {freq} ==", flush=True)
    collapse = rows_to_daily if is_daily else rows_to_monthly

    from bbg import BBG
    _setup_with_retry()
    series = {}
    t0 = _time.time()
    with BBG() as bbg:
        print("  probe:", bbg.ref("AAPL US Equity", "PX_LAST"), flush=True)
        for fi, (fld, tickers) in enumerate(sorted(by_field.items()), 1):
            tickers = sorted(tickers)
            try:
                batch = bbg.hist_batch(tickers, fld, start, end, periodicity=freq,
                                       currency=currency, calendar_fill=is_daily)
            except Exception as e:
                print(f"  WARN batch failed field={fld}: {e}", flush=True)
                batch = {}
            for tk in tickers:
                series[(tk, fld)] = collapse(batch.get(tk, []), fld)
            print(f"    [{fi}/{len(by_field)}] {fld}: {len(tickers)} tickers ({_time.time()-t0:.1f}s)", flush=True)
    print(f"  all pulls done in {_time.time()-t0:.1f}s", flush=True)

    # assemble each direct sheet: date index = union of its columns' dates
    sheet_frames = {}
    for s in sheets:
        cols = s["columns"]
        all_idx = pd.DatetimeIndex([]) if is_daily else pd.PeriodIndex([], freq="M")
        for c in cols:
            sr = series.get((c["ticker"], c["field"]), pd.Series(dtype=float))
            all_idx = all_idx.union(sr.index)
        all_idx = all_idx.sort_values()
        frame = pd.DataFrame(index=all_idx)
        meta = {}  # col -> (ticker, field)
        for c in cols:
            frame[c["col"]] = series.get((c["ticker"], c["field"]), pd.Series(dtype=float)).reindex(all_idx).ffill()  # Fill=P
            meta[c["col"]] = (c["ticker"], c["field"])
        if is_daily:
            # keep every complete trading day up to and including the last available
            frame = frame[frame.index <= pd.Timestamp(date.today())]
            frame.index = frame.index.normalize()
        else:
            # drop the current (incomplete) month — dump ends at last COMPLETE month.
            frame = frame[frame.index < pd.Period(date.today(), freq="M")]
            # month-period -> month-end timestamp for output
            frame.index = frame.index.to_timestamp(how="end").normalize()
        sheet_frames[s["sheet"]] = (frame, meta)

    # calc sheets (only if MCAP present)
    if "MCAP" in sheet_frames and not only:
        mcap_frame, mcap_meta = sheet_frames["MCAP"]
        factors = manifest["calc"]["mcap_split_factors"]   # 34, aligned to MCAP columns (col2..col35)
        cols_sorted = sorted(mcap_frame.columns)
        adj = mcap_frame.copy()
        for j, col in enumerate(cols_sorted):
            adj[col] = mcap_frame[col] * float(factors[j])
        sheet_frames["MCAP Adj"] = (adj, mcap_meta)
        wts = adj.div(adj.sum(axis=1), axis=0)
        sheet_frames["Mcap Weights"] = (wts, mcap_meta)

    # write workbook in manifest sheet order, Master first as a static placeholder
    out = out_path
    out.parent.mkdir(parents=True, exist_ok=True)
    order = ["Master"] + [s["sheet"] for s in sheets] + (CALC_SHEETS if not only else [])
    # de-dup order preserving first occurrence
    seen = set(); order = [x for x in order if not (x in seen or seen.add(x))]

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    # Master placeholder (build_t2_master skips sheet 0)
    ws0 = wb.create_sheet("Master")
    ws0["A1"] = "ETF"
    for r, etf in enumerate(manifest["calc"]["etf_order"], start=2):
        ws0.cell(r, 1, etf)
    for c, fac in enumerate(manifest["calc"]["mcap_split_factors"], start=2):
        ws0.cell(39, c, fac)

    for name in order:
        if name == "Master":
            continue
        if name not in sheet_frames:
            continue
        frame, meta = sheet_frames[name]
        ws = wb.create_sheet(name)
        cols_sorted = sorted(frame.columns)
        ws.cell(1, 1, None); ws.cell(2, 1, "Dates")
        for k, col in enumerate(cols_sorted, start=2):
            tk, fld = meta.get(col, (None, None))
            ws.cell(1, k, tk)
            ws.cell(2, k, fld if fld else "PX_LAST")
        for r, (dt, row) in enumerate(frame.iterrows(), start=3):
            ws.cell(r, 1, pd.Timestamp(dt).to_pydatetime())
            for k, col in enumerate(cols_sorted, start=2):
                v = row[col]
                ws.cell(r, k, None if pd.isna(v) else float(v))

    wb.save(out)
    print(f"\n== wrote {out} ({len(wb.sheetnames)} sheets) ==", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
