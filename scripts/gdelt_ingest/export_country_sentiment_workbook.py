#!/usr/bin/env python3
"""
=============================================================================
SCRIPT NAME: export_country_sentiment_workbook.py
=============================================================================

INPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/data/aggregates/*/country_day_all.csv
    Per-month aggregate CSV files containing country-day GDELT data
    (used when --panel-csv and --panel-parquet are not provided)
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/data/panels/<panel>.parquet
    Prebuilt panel parquet file (preferred input, provided via --panel-parquet)
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/data/panels/<panel>.csv
    Prebuilt panel CSV file (provided via --panel-csv)

OUTPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/output/spreadsheet/gdelt_country_sentiment.xlsx
    Multi-sheet Excel workbook with GDELT country sentiment indicators.
    One sheet per indicator, README, and README_VARIABLES sheets.
    Layout: dates down column A, country buckets across row 1.

VERSION: 1.0
LAST UPDATED: 2026-06-05
AUTHOR: Arjun Divecha

DESCRIPTION:
Export GDELT country sentiment indicators into a multi-sheet Excel workbook.
Supports both monthly and daily panels. Includes country bucket aliases
(e.g. U.S., U.S. NASDAQ, US SmallCap all map to USA), variable definitions,
and template-based styling.

DEPENDENCIES:
- pandas
- openpyxl
- Python 3.x

USAGE:
    python export_country_sentiment_workbook.py
    python export_country_sentiment_workbook.py --panel-parquet <path> --output <path>

NOTES:
- Default paths are relative to the GDELT project root.
- COUNTRY_BUCKETS, INDICATORS, and VARIABLE_DEFINITIONS are imported by
  export_daily_workbook.py (same directory).
- An optional template workbook can be provided via --template-xlsx to borrow styles.
- Incomplete months are automatically filtered out for monthly panels.
=============================================================================
"""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


COUNTRY_BUCKETS = [
    ("Singapore", "SGP"),
    ("Australia", "AUS"),
    ("Canada", "CAN"),
    ("Germany", "DEU"),
    ("Japan", "JPN"),
    ("Switzerland", "CHE"),
    ("U.K.", "GBR"),
    ("U.S. NASDAQ", "USA"),
    ("U.S.", "USA"),
    ("France", "FRA"),
    ("Netherlands", "NLD"),
    ("Sweden", "SWE"),
    ("Italy", "ITA"),
    ("China A", "CHN"),
    ("Chile", "CHL"),
    ("Indonesia", "IDN"),
    ("Philippines", "PHL"),
    ("Poland", "POL"),
    ("US SmallCap", "USA"),
    ("Malaysia", "MYS"),
    ("Taiwan", "TWN"),
    ("Mexico", "MEX"),
    ("Korea", "KOR"),
    ("Brazil", "BRA"),
    ("South Africa", "ZAF"),
    ("Denmark", "DNK"),
    ("India", "IND"),
    ("China H", "CHN"),
    ("Hong Kong", "HKG"),
    ("Thailand", "THA"),
    ("Turkey", "TUR"),
    ("Spain", "ESP"),
    ("Vietnam", "VNM"),
    ("Saudi Arabia", "SAU"),
]

INDICATORS = [
    ("monthly_metronome", "monthly_metronome"),
    ("monthly_risk", "monthly_risk"),
    ("monthly_defensive", "monthly_defensive"),
    ("metronome_rank_pct", "monthly_metronome_rank_pct"),
    ("risk_rank_pct", "monthly_risk_rank_pct"),
    ("defensive_rank_pct", "monthly_defensive_rank_pct"),
    ("sentiment_fast", "sentiment_fast"),
    ("sentiment_slow", "sentiment_slow"),
    ("sentiment_trend", "sentiment_trend"),
    ("attention_fast", "attention_fast"),
    ("attention_slow", "attention_slow"),
    ("attention_trend", "attention_trend"),
    ("risk_fast", "risk_fast"),
    ("dispersion_fast", "dispersion_fast"),
    ("local_tone_fast", "local_tone_fast"),
    ("foreign_tone_fast", "foreign_tone_fast"),
    ("local_foreign_gap", "local_foreign_gap"),
    ("sentiment_fast_z", "sentiment_fast_z"),
    ("sentiment_slow_z", "sentiment_slow_z"),
    ("sentiment_trend_z", "sentiment_trend_z"),
    ("attention_fast_z", "attention_fast_z"),
    ("attention_slow_z", "attention_slow_z"),
    ("attention_trend_z", "attention_trend_z"),
    ("risk_fast_z", "risk_fast_z"),
    ("dispersion_fast_z", "dispersion_fast_z"),
    ("local_tone_fast_z", "local_tone_fast_z"),
    ("foreign_tone_fast_z", "foreign_tone_fast_z"),
    ("lf_gap_z", "local_foreign_gap_z"),
    ("country_news_sentiment", "country_news_sentiment"),
    ("country_news_risk", "country_news_risk"),
    ("country_news_sentiment_raw", "country_news_sentiment_raw"),
    ("country_news_risk_raw", "country_news_risk_raw"),
    ("country_news_attention", "country_news_attention"),
    ("local_attention_share", "local_attention_share"),
    ("sentiment_x_attention", "country_news_sentiment_x_attention"),
    ("local_tone", "local_tone"),
    ("foreign_tone", "foreign_tone"),
    ("attention_shock", "attention_shock"),
    ("tone_dispersion", "tone_dispersion"),
    ("tone_wavg_wordcount", "tone_wavg_wordcount"),
    ("tone_mean", "tone_mean"),
    ("tone_p50", "tone_p50"),
    ("positive_mean", "positive_mean"),
    ("negative_mean", "negative_mean"),
    ("polarity_mean", "polarity_mean"),
    ("n_articles", "n_articles"),
]


VARIABLE_DEFINITIONS = {
    "monthly_metronome": (
        "Primary monthly composite signal. Higher values indicate a stronger positive signal after accounting for risk.",
        "0.35*sentiment_fast_z + 0.20*sentiment_slow_z + 0.20*sentiment_trend_z + 0.15*attention_fast_z - 0.10*risk_fast_z",
        "monthly",
    ),
    "monthly_risk": (
        "Primary monthly risk composite. Higher values indicate more risk-off conditions.",
        "0.45*risk_fast_z + 0.30*dispersion_fast_z - 0.15*sentiment_fast_z - 0.10*foreign_tone_fast_z",
        "monthly",
    ),
    "monthly_defensive": (
        "Defensive positioning score. Positive values indicate more defensive conditions.",
        "-1.0 * monthly_risk",
        "monthly",
    ),
    "monthly_metronome_rank_pct": (
        "Cross-country percentile rank of monthly_metronome within each month (0 to 1).",
        "groupby(signal_month).rank(pct=True)",
        "monthly",
    ),
    "monthly_risk_rank_pct": (
        "Cross-country percentile rank of monthly_risk within each month (0 to 1).",
        "groupby(signal_month).rank(pct=True)",
        "monthly",
    ),
    "monthly_defensive_rank_pct": (
        "Cross-country percentile rank of monthly_defensive within each month (0 to 1).",
        "groupby(signal_month).rank(pct=True)",
        "monthly",
    ),
    "sentiment_fast": (
        "Fast exponentially weighted monthly sentiment level.",
        "EWMA(span=5) of country_news_sentiment_raw",
        "monthly",
    ),
    "sentiment_slow": (
        "Slow exponentially weighted monthly sentiment level.",
        "EWMA(span=20) of country_news_sentiment_raw",
        "monthly",
    ),
    "sentiment_trend": (
        "Sentiment momentum signal.",
        "sentiment_fast - sentiment_slow",
        "monthly",
    ),
    "attention_fast": (
        "Fast exponentially weighted monthly local attention share.",
        "EWMA(span=5) of local_attention_share",
        "monthly",
    ),
    "attention_slow": (
        "Slow exponentially weighted monthly local attention share.",
        "EWMA(span=20) of local_attention_share",
        "monthly",
    ),
    "attention_trend": (
        "Attention momentum signal.",
        "attention_fast - attention_slow",
        "monthly",
    ),
    "risk_fast": (
        "Fast exponentially weighted monthly risk level.",
        "EWMA(span=10) of country_news_risk_raw",
        "monthly",
    ),
    "dispersion_fast": (
        "Fast exponentially weighted monthly tone dispersion.",
        "EWMA(span=10) of tone_dispersion",
        "monthly",
    ),
    "local_tone_fast": (
        "Fast exponentially weighted local-source tone.",
        "EWMA(span=5) of local_tone",
        "monthly",
    ),
    "foreign_tone_fast": (
        "Fast exponentially weighted foreign-source tone.",
        "EWMA(span=5) of foreign_tone",
        "monthly",
    ),
    "local_foreign_gap": (
        "Difference between local and foreign tone in the fast block.",
        "local_tone_fast - foreign_tone_fast",
        "monthly",
    ),
    "sentiment_fast_z": (
        "Within-country standardized version of sentiment_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "sentiment_slow_z": (
        "Within-country standardized version of sentiment_slow using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "sentiment_trend_z": (
        "Within-country standardized version of sentiment_trend using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "attention_fast_z": (
        "Within-country standardized version of attention_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "attention_slow_z": (
        "Within-country standardized version of attention_slow using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "attention_trend_z": (
        "Within-country standardized version of attention_trend using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "risk_fast_z": (
        "Within-country standardized version of risk_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "dispersion_fast_z": (
        "Within-country standardized version of dispersion_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "local_tone_fast_z": (
        "Within-country standardized version of local_tone_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "foreign_tone_fast_z": (
        "Within-country standardized version of foreign_tone_fast using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "local_foreign_gap_z": (
        "Within-country standardized version of local_foreign_gap using trailing monthly history.",
        "trailing z-score (window=24 months, min_history=6, shift(1))",
        "monthly",
    ),
    "country_news_sentiment": (
        "Daily standardized country sentiment snapshot at month-end.",
        "daily trailing z-score of country_news_sentiment_raw",
        "daily->monthly snapshot",
    ),
    "country_news_risk": (
        "Daily standardized country risk snapshot at month-end.",
        "daily trailing z-score of country_news_risk_raw",
        "daily->monthly snapshot",
    ),
    "country_news_sentiment_raw": (
        "Base raw sentiment level used by downstream composites.",
        "local_tone (fallback to tone_wavg_wordcount when local_tone is missing in daily build)",
        "daily",
    ),
    "country_news_risk_raw": (
        "Base raw risk level used by downstream composites.",
        "-country_news_sentiment_raw + 0.5*tone_dispersion",
        "daily",
    ),
    "country_news_attention": (
        "Log-scaled local article attention signal.",
        "log(1 + local_n_articles)",
        "daily",
    ),
    "local_attention_share": (
        "Share of local-source article attention.",
        "local_n_articles / local_source_total_articles",
        "daily",
    ),
    "country_news_sentiment_x_attention": (
        "Standardized interaction between sentiment and local attention share.",
        "daily trailing z-score of (country_news_sentiment_raw * local_attention_share)",
        "daily",
    ),
    "local_tone": (
        "Average tone of local-source-attributed articles.",
        "from daily aggregate inputs",
        "daily",
    ),
    "foreign_tone": (
        "Average tone of foreign-source-attributed articles.",
        "from daily aggregate inputs",
        "daily",
    ),
    "attention_shock": (
        "Standardized attention spike indicator.",
        "daily trailing z-score of country_news_attention",
        "daily",
    ),
    "tone_dispersion": (
        "Spread of tone values; higher means less agreement in tone.",
        "from daily aggregate inputs",
        "daily",
    ),
    "tone_wavg_wordcount": (
        "Word-count-weighted average tone.",
        "from daily aggregate inputs",
        "daily",
    ),
    "tone_mean": (
        "Simple average tone.",
        "from daily aggregate inputs",
        "daily",
    ),
    "negative_mean": (
        "Average negative-tone component.",
        "from daily aggregate inputs",
        "daily",
    ),
    "n_articles": (
        "Total article count for the country-day snapshot used at month-end.",
        "from daily aggregate inputs",
        "daily",
    ),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export GDELT country-day aggregates into a workbook in the sample country layout."
    )
    parser.add_argument(
        "--aggregates-dir",
        default="data/aggregates",
        help="Directory containing daily aggregate folders with country_day_all.csv files.",
    )
    parser.add_argument(
        "--panel-csv",
        default="",
        help="Optional prebuilt panel CSV. If provided, this is used instead of data/aggregates/*/country_day_all.csv.",
    )
    parser.add_argument(
        "--panel-parquet",
        default="",
        help="Optional prebuilt panel parquet. Preferred for the stream-first pipeline.",
    )
    parser.add_argument(
        "--output",
        default="output/spreadsheet/gdelt_country_sentiment.xlsx",
        help="Path to the output workbook.",
    )
    parser.add_argument(
        "--template-xlsx",
        default="",
        help="Optional workbook to borrow styles from.",
    )
    return parser.parse_args()


def load_panel(aggregates_dir: Path, panel_csv: str, panel_parquet: str) -> pd.DataFrame:
    if panel_parquet:
        frame = pd.read_parquet(panel_parquet)
    elif panel_csv:
        frame = pd.read_csv(panel_csv)
    else:
        parquet_files = sorted(aggregates_dir.glob("*.parquet"))
        if parquet_files:
            frame = pd.concat((pd.read_parquet(path) for path in parquet_files), ignore_index=True)
        else:
            files = sorted(aggregates_dir.glob("*/country_day_all.csv"))
            if not files:
                raise FileNotFoundError(
                    f"No country_day_all.csv files or parquet panel files found under {aggregates_dir}"
                )
            frame = pd.concat((pd.read_csv(path) for path in files), ignore_index=True)
    frame = frame.dropna(subset=["country_iso3"]).copy()
    frame["country_iso3"] = frame["country_iso3"].astype(str).str.strip()
    frame = frame.loc[frame["country_iso3"] != ""].copy()
    if "signal_month_end_date" in frame.columns:
        frame["row_date"] = pd.to_datetime(frame["signal_month_end_date"]).dt.normalize()
    elif "signal_month" in frame.columns:
        # Monthly metronome panels may carry signal_month without an explicit
        # calendar month-end date column; normalize these to month-end rows.
        frame["row_date"] = (
            pd.to_datetime(frame["signal_month"].astype(str), errors="coerce")
            .dt.to_period("M")
            .dt.to_timestamp(how="end")
            .dt.normalize()
        )
    else:
        frame["row_date"] = pd.to_datetime(frame["date"]).dt.normalize()
    frame["date"] = pd.to_datetime(frame["date"]).dt.normalize()
    if "signal_month" in frame.columns:
        # Monthly panels can include an in-progress current month (for example,
        # March rows built from data through March 4 but labeled 2026-03-31).
        # Keep only fully completed months in workbook exports.
        max_observed_date = frame["date"].max()
        max_month_end = max_observed_date.to_period("M").to_timestamp(how="end").normalize()
        if max_observed_date < max_month_end:
            completed_month_end = (
                (max_observed_date.to_period("M") - 1).to_timestamp(how="end").normalize()
            )
            frame = frame.loc[frame["row_date"] <= completed_month_end].copy()
    frame = frame.sort_values(["row_date", "country_iso3", "date"]).drop_duplicates(
        subset=["row_date", "country_iso3"], keep="last"
    )
    return frame


def make_style_book(template_path: str) -> tuple[Workbook, dict[str, object]]:
    if template_path:
        template_wb = load_workbook(template_path)
        template_ws = template_wb[template_wb.sheetnames[0]]
        style_refs = {
            "header_style": copy(template_ws["B1"]._style),
            "date_style": copy(template_ws["A2"]._style),
            "value_style": copy(template_ws["B2"]._style),
            "header_font": copy(template_ws["B1"].font),
            "date_font": copy(template_ws["A2"].font),
            "value_font": copy(template_ws["B2"].font),
            "header_fill": copy(template_ws["B1"].fill),
            "date_fill": copy(template_ws["A2"].fill),
            "value_fill": copy(template_ws["B2"].fill),
            "header_border": copy(template_ws["B1"].border),
            "date_border": copy(template_ws["A2"].border),
            "value_border": copy(template_ws["B2"].border),
            "header_alignment": copy(template_ws["B1"].alignment),
            "date_alignment": copy(template_ws["A2"].alignment),
            "value_alignment": copy(template_ws["B2"].alignment),
            "header_number_format": template_ws["B1"].number_format,
            "date_number_format": template_ws["A2"].number_format,
            "value_number_format": template_ws["B2"].number_format,
            "column_a_width": template_ws.column_dimensions["A"].width,
        }
        wb = Workbook()
        wb.remove(wb.active)
        return wb, style_refs

    wb = Workbook()
    wb.remove(wb.active)
    style_refs = {
        "column_a_width": 10.33,
        "date_number_format": "mm-dd-yy",
        "value_number_format": "0.000000",
    }
    return wb, style_refs


def apply_template_style(cell, prefix: str, style_refs: dict[str, object]) -> None:
    style_key = f"{prefix}_style"
    if style_key in style_refs:
        cell._style = copy(style_refs[style_key])
        cell.font = copy(style_refs[f"{prefix}_font"])
        cell.fill = copy(style_refs[f"{prefix}_fill"])
        cell.border = copy(style_refs[f"{prefix}_border"])
        cell.alignment = copy(style_refs[f"{prefix}_alignment"])
        cell.number_format = style_refs[f"{prefix}_number_format"]


def populate_sheet(
    ws, wide: pd.DataFrame, sheet_name: str, indicator: str, style_refs: dict[str, object]
) -> None:
    ws.title = sheet_name
    ws.column_dimensions["A"].width = style_refs["column_a_width"]

    ws.cell(row=1, column=1, value=None)
    for idx, (label, _iso3) in enumerate(COUNTRY_BUCKETS, start=2):
        cell = ws.cell(row=1, column=idx, value=label)
        apply_template_style(cell, "header", style_refs)

    for row_idx, dt in enumerate(wide.index, start=2):
        date_cell = ws.cell(row=row_idx, column=1, value=dt.to_pydatetime())
        apply_template_style(date_cell, "date", style_refs)
        if "date_number_format" in style_refs:
            date_cell.number_format = style_refs["date_number_format"]

        for col_idx, (label, iso3) in enumerate(COUNTRY_BUCKETS, start=2):
            value = wide.at[dt, iso3] if iso3 in wide.columns else None
            if pd.isna(value):
                value = None
            value_cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_template_style(value_cell, "value", style_refs)
            if "value_number_format" in style_refs and indicator != "n_articles":
                value_cell.number_format = style_refs["value_number_format"]
            if indicator == "n_articles":
                value_cell.number_format = "0"


def build_indicator_panel(frame: pd.DataFrame, indicator: str) -> pd.DataFrame:
    panel = frame.pivot(index="row_date", columns="country_iso3", values=indicator)
    panel = panel.sort_index()
    needed = sorted({iso3 for _label, iso3 in COUNTRY_BUCKETS})
    panel = panel.reindex(columns=needed)
    return panel


def add_readme_sheet(wb: Workbook, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet("README", 0)
    min_date = frame["date"].min().date().isoformat()
    max_date = frame["date"].max().date().isoformat()
    available_indicators = [sheet_name for sheet_name, indicator in INDICATORS if indicator in frame.columns]
    lines = [
        "GDELT country sentiment workbook",
        f"Coverage in workbook: {min_date} to {max_date}",
        "Layout: dates down column A, country buckets across row 1.",
        "Aliases: U.S., U.S. NASDAQ, and US SmallCap all map to USA.",
        "Aliases: China A and China H both map to CHN.",
        "Source data: panel parquet/CSV or data/aggregates/*/country_day_all.csv",
        "Workbook sheets:",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    start_row = len(lines) + 1
    for offset, sheet_name in enumerate(available_indicators, start=start_row):
        ws.cell(row=offset, column=1, value=sheet_name)


def add_variable_dictionary_sheet(wb: Workbook, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet("README_VARIABLES", 1)
    headers = ["sheet_name", "pipeline_column", "stage", "definition", "formula_or_source"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row = 2
    for sheet_name, indicator in INDICATORS:
        if indicator not in frame.columns:
            continue

        pipeline_column = indicator
        if sheet_name == "lf_gap_z":
            pipeline_column = "local_foreign_gap_z"
        elif sheet_name == "sentiment_x_attention":
            pipeline_column = "country_news_sentiment_x_attention"

        definition, formula, stage = VARIABLE_DEFINITIONS.get(
            pipeline_column,
            (
                "No detailed definition found in VARIABLE_DEFINITIONS.",
                "See source pipeline scripts for implementation details.",
                "unknown",
            ),
        )

        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=pipeline_column)
        ws.cell(row=row, column=3, value=stage)
        ws.cell(row=row, column=4, value=definition)
        ws.cell(row=row, column=5, value=formula)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 84
    ws.column_dimensions["E"].width = 84


def main() -> None:
    args = parse_args()
    aggregates_dir = Path(args.aggregates_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    frame = load_panel(aggregates_dir, args.panel_csv, args.panel_parquet)
    wb, style_refs = make_style_book(args.template_xlsx)
    add_readme_sheet(wb, frame)
    add_variable_dictionary_sheet(wb, frame)

    for sheet_name, indicator in INDICATORS:
        if indicator not in frame.columns:
            continue
        wide = build_indicator_panel(frame, indicator)
        ws = wb.create_sheet(sheet_name)
        populate_sheet(ws, wide, sheet_name, indicator, style_refs)

    wb.save(output_path)
    print(f"saved {output_path}")
    print(f"rows={frame['row_date'].nunique()} dates, countries={len(COUNTRY_BUCKETS)} buckets")


if __name__ == "__main__":
    main()
