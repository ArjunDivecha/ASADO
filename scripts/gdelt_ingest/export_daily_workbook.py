#!/usr/bin/env python3
"""
=============================================================================
SCRIPT NAME: export_daily_workbook.py
=============================================================================

INPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/data/panels/country_signal_daily_metronome.parquet
    Daily metronome panel parquet (default for --panel-parquet)

OUTPUT FILES:
- /Users/arjundivecha/Dropbox/AAA Backup/A Complete/GDELT/output/spreadsheet/GDELT_DAILY.xlsx
    Multi-sheet Excel workbook with daily GDELT metronome indicators
    (default for --output)

VERSION: 1.0
LAST UPDATED: 2026-06-05
AUTHOR: Arjun Divecha

DESCRIPTION:
Export the daily GDELT metronome panel into a multi-sheet Excel workbook.
Layout: dates (daily) down column A, country buckets across row 1.
One sheet per indicator. This is the daily analog of
export_country_sentiment_workbook.py.

DEPENDENCIES:
- pandas
- openpyxl
- pathlib
- Python 3.x

USAGE:
    python export_daily_workbook.py
    python export_daily_workbook.py --panel-parquet <path> --output <path>

NOTES:
- Default paths are relative to the GDELT project root.
- Imports COUNTRY_BUCKETS, apply_template_style, make_style_book,
  and populate_sheet from export_country_sentiment_workbook.py (same dir).
=============================================================================
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Reuse shared infrastructure from the monthly exporter
# ---------------------------------------------------------------------------
import sys

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from export_country_sentiment_workbook import (  # noqa: E402
    COUNTRY_BUCKETS,
    apply_template_style,
    make_style_book,
    populate_sheet,
)


# ---------------------------------------------------------------------------
# Daily indicator list
# ---------------------------------------------------------------------------

DAILY_INDICATORS = [
    # ── Composites ──
    ("daily_metronome", "daily_metronome"),
    ("daily_risk", "daily_risk"),
    ("daily_defensive", "daily_defensive"),
    # ── Cross-sectional ranks ──
    ("metronome_rank_pct", "daily_metronome_rank_pct"),
    ("risk_rank_pct", "daily_risk_rank_pct"),
    ("defensive_rank_pct", "daily_defensive_rank_pct"),
    # ── Raw EWMA features ──
    ("sentiment_fast", "sentiment_fast"),
    ("sentiment_slow", "sentiment_slow"),
    ("sentiment_trend", "sentiment_trend"),
    ("attention_fast", "attention_fast"),
    ("attention_slow", "attention_slow"),
    ("attention_trend", "attention_trend"),
    ("risk_fast", "risk_fast"),
    ("dispersion_fast", "dispersion_fast"),
    ("local_tone_fast", "local_tone_fast"),
    ("foreign_tone_fast", "foreign_tone_fast"),
    ("local_foreign_gap", "local_foreign_gap"),
    # ── Z-scored EWMA features ──
    ("sentiment_fast_z", "sentiment_fast_z"),
    ("sentiment_slow_z", "sentiment_slow_z"),
    ("sentiment_trend_z", "sentiment_trend_z"),
    ("attention_fast_z", "attention_fast_z"),
    ("attention_slow_z", "attention_slow_z"),
    ("attention_trend_z", "attention_trend_z"),
    ("risk_fast_z", "risk_fast_z"),
    ("dispersion_fast_z", "dispersion_fast_z"),
    ("local_tone_fast_z", "local_tone_fast_z"),
    ("foreign_tone_fast_z", "foreign_tone_fast_z"),
    ("lf_gap_z", "local_foreign_gap_z"),
    # ── Underlying daily raw signals ──
    ("country_news_sentiment_raw", "country_news_sentiment_raw"),
    ("country_news_risk_raw", "country_news_risk_raw"),
    ("country_news_attention", "country_news_attention"),
    ("local_attention_share", "local_attention_share"),
    ("local_tone", "local_tone"),
    ("foreign_tone", "foreign_tone"),
    ("attention_shock", "attention_shock"),
    ("tone_dispersion", "tone_dispersion"),
    ("n_articles", "n_articles"),
]

DAILY_VARIABLE_DEFINITIONS = {
    "daily_metronome": (
        "Primary daily composite signal.  Identical formula to monthly_metronome "
        "but z-scored over a trailing 504-day window instead of 24 months.",
        "0.35*sentiment_fast_z + 0.20*sentiment_slow_z + 0.20*sentiment_trend_z "
        "+ 0.15*attention_fast_z - 0.10*risk_fast_z",
        "daily",
    ),
    "daily_risk": (
        "Primary daily risk composite.  Identical formula to monthly_risk "
        "but z-scored over a trailing 504-day window.",
        "0.45*risk_fast_z + 0.30*dispersion_fast_z - 0.15*sentiment_fast_z "
        "- 0.10*foreign_tone_fast_z",
        "daily",
    ),
    "daily_defensive": (
        "Defensive positioning score (daily). Positive = more defensive.",
        "-1.0 * daily_risk",
        "daily",
    ),
    "daily_metronome_rank_pct": (
        "Cross-country percentile rank of daily_metronome within each day (0 to 1).",
        "groupby(date).rank(pct=True)",
        "daily",
    ),
    "daily_risk_rank_pct": (
        "Cross-country percentile rank of daily_risk within each day (0 to 1).",
        "groupby(date).rank(pct=True)",
        "daily",
    ),
    "daily_defensive_rank_pct": (
        "Cross-country percentile rank of daily_defensive within each day (0 to 1).",
        "groupby(date).rank(pct=True)",
        "daily",
    ),
}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export the daily GDELT metronome panel into a multi-sheet Excel workbook."
    )
    parser.add_argument(
        "--panel-parquet",
        default="data/panels/country_signal_daily_metronome.parquet",
        help="Input daily metronome parquet.",
    )
    parser.add_argument(
        "--output",
        default="output/spreadsheet/GDELT_DAILY.xlsx",
        help="Path to the output workbook.",
    )
    parser.add_argument(
        "--template-xlsx",
        default="",
        help="Optional template workbook to borrow styles from.",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Panel loading
# ---------------------------------------------------------------------------

def load_daily_panel(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Daily metronome panel not found: {path}")

    frame = pd.read_parquet(path)
    frame = frame.dropna(subset=["country_iso3"]).copy()
    frame["country_iso3"] = frame["country_iso3"].astype(str).str.strip()
    frame = frame.loc[frame["country_iso3"] != ""].copy()
    frame["date"] = pd.to_datetime(frame["date"]).dt.normalize()

    # row_date is just the actual observation date (no month-end shifting)
    frame["row_date"] = frame["date"]

    frame = frame.sort_values(["row_date", "country_iso3"]).drop_duplicates(
        subset=["row_date", "country_iso3"], keep="last"
    )
    return frame


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def build_indicator_panel(frame: pd.DataFrame, indicator: str) -> pd.DataFrame:
    """Pivot to wide: dates × countries."""
    panel = frame.pivot(index="row_date", columns="country_iso3", values=indicator)
    panel = panel.sort_index()
    needed = sorted({iso3 for _label, iso3 in COUNTRY_BUCKETS})
    panel = panel.reindex(columns=needed)
    return panel


def add_readme_sheet(wb: Workbook, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet("README", 0)
    min_date = frame["date"].min().date().isoformat()
    max_date = frame["date"].max().date().isoformat()
    available = [sn for sn, ind in DAILY_INDICATORS if ind in frame.columns]
    lines = [
        "GDELT daily country sentiment workbook",
        f"Coverage: {min_date} to {max_date}",
        "Frequency: daily (every calendar day with GDELT data)",
        "Layout: dates down column A, country buckets across row 1.",
        "Aliases: U.S., U.S. NASDAQ, and US SmallCap all map to USA.",
        "Aliases: China A and China H both map to CHN.",
        "Z-score window: 504 calendar days (~24 months), min history 126 days (~6 months).",
        "Composite formulae are identical to the monthly version.",
        "",
        "Workbook sheets:",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    start_row = len(lines) + 1
    for offset, sheet_name in enumerate(available, start=start_row):
        ws.cell(row=offset, column=1, value=sheet_name)


def add_variable_dictionary_sheet(wb: Workbook, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet("README_VARIABLES", 1)
    headers = ["sheet_name", "pipeline_column", "stage", "definition", "formula_or_source"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Import monthly definitions as fallback
    from export_country_sentiment_workbook import VARIABLE_DEFINITIONS as MONTHLY_DEFS

    row = 2
    for sheet_name, indicator in DAILY_INDICATORS:
        if indicator not in frame.columns:
            continue

        pipeline_column = indicator
        if sheet_name == "lf_gap_z":
            pipeline_column = "local_foreign_gap_z"

        definition, formula, stage = DAILY_VARIABLE_DEFINITIONS.get(
            pipeline_column,
            MONTHLY_DEFS.get(
                pipeline_column,
                (
                    "See source pipeline scripts for details.",
                    "—",
                    "daily",
                ),
            ),
        )

        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=pipeline_column)
        ws.cell(row=row, column=3, value=stage)
        ws.cell(row=row, column=4, value=definition)
        ws.cell(row=row, column=5, value=formula)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 84
    ws.column_dimensions["E"].width = 84


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_daily_workbook(
    frame: pd.DataFrame,
    output_path: Path,
    template_xlsx: str = "",
) -> None:
    """Export the daily metronome panel to a multi-sheet Excel workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb, style_refs = make_style_book(template_xlsx)

    add_readme_sheet(wb, frame)
    add_variable_dictionary_sheet(wb, frame)

    sheet_count = 0
    for sheet_name, indicator in DAILY_INDICATORS:
        if indicator not in frame.columns:
            continue
        wide = build_indicator_panel(frame, indicator)
        ws = wb.create_sheet(sheet_name)
        populate_sheet(ws, wide, sheet_name, indicator, style_refs)
        sheet_count += 1

    wb.save(output_path)
    print(
        f"saved {output_path} "
        f"({sheet_count} sheets, "
        f"{frame['row_date'].nunique()} daily dates, "
        f"{len(COUNTRY_BUCKETS)} country buckets)"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()
    panel_path = Path(args.panel_parquet)
    output_path = Path(args.output)

    frame = load_daily_panel(panel_path)
    export_daily_workbook(frame, output_path, args.template_xlsx)


if __name__ == "__main__":
    main()
